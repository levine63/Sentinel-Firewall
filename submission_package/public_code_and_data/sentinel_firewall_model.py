"""Sentinel Firewall safe-products economic model.

Purpose
-------
This script runs the probabilistic sensitivity analysis for the manuscript
"The Sentinel Firewall: A Health-System Strategy to Decouple Informal
Industrialization from Lead Exposure in LMICs." It compares two modular
lead-exposure reduction packages:

1. Kitchen Package: lead-safe pot voucher, calcium, and safe child utensils.
2. Kohl Package: safe maternal kohl and safe infant kohl.

The model estimates benefit-cost ratios from child cognition and lifetime
earnings, then adds maternal/neonatal health benefits and adult cardiovascular
benefits as secondary extensions.

How to run
----------
Install Python 3.11+ and the required packages:

    pip install numpy pandas scipy matplotlib openpyxl

Then run:

    python sentinel_firewall_model.py

The script creates an ``outputs`` folder containing trial-level CSV files,
summary tables, validation checks, figure files, and an editable deterministic
Excel workbook.

How to modify parameters
------------------------
Most model inputs are embedded in the ``CSV_DATA`` block below. Each row has:

    variable, min, mode, max, is_cost, description, source_note

To adapt the model to a specific country or district, edit the min/mode/max
values in ``CSV_DATA`` and rerun the script. The deterministic spreadsheet
``Simulating_sentinel_safe_products.xlsx`` is easier for modifying only modal
values, but this Python script is the authoritative source for Monte Carlo
uncertainty, correlations, plots, and trial-level exports.

Important modelling notes
-------------------------
- PERT distributions are used for uncertain parameters.
- Some implementation steps are positively correlated to reflect stronger and
  weaker local delivery environments.
- Maternal lead reductions can reduce fetal lead exposure through an explicit
  fetal-transfer parameter.
- Child cognition benefits are developmentally weighted so short interventions
  do not count as if they changed exposure across the full early-childhood
  window.
- Adult cardiovascular benefits are reported as a secondary extension.

Outputs to cite or inspect
--------------------------
- outputs/summary_table.csv
- outputs/validation_checks.csv
- outputs/kitchen_baseline_trials.csv.gz
- outputs/kohl_baseline_trials.csv.gz
- outputs/figure_1_bcr_distribution.png
- outputs/Simulating_sentinel_safe_products.xlsx

Author/contact
--------------
David I. Levine, University of California, Berkeley.
Correspondence: levine@berkeley.edu
ORCID: 0000-0002-7210-5499
"""

import io
from pathlib import Path
import shutil

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Protection
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from scipy import stats
from scipy.linalg import cholesky

try:
    import matplotlib.pyplot as plt
    from matplotlib.patches import FancyArrowPatch, FancyBboxPatch
except ModuleNotFoundError:
    plt = None
    FancyArrowPatch = None
    FancyBboxPatch = None


N_TRIALS = 10000
RNG_SEED = 42
NONLINEAR_BETA = 3.246
TORNADO_TOP_N = 12


CSV_DATA = """variable,min,mode,max,is_cost,description,source_note
gdp_ppp_per_capita,2000,4000,8000,FALSE,PPP GDP per capita,World Bank LMIC context
labor_share_income,0.5,0.55,0.58,FALSE,Labour share of national income including corrected self-employment income,ILO labour-income-share data with Gollin self-employment correction
growth,0.0,0.02,0.025,FALSE,Real annual productivity growth,Long-run LMIC productivity assumption
discount,0.01,0.03,0.05,TRUE,Social discount rate,Standard economic evaluation range
p18,0.92,0.95,0.97,FALSE,Probability survive to age 18,World Bank or UN life-table context
p65,0.68,0.78,0.85,FALSE,Probability survive to age 65,World Bank or UN life-table context
k,1.5,2.0,3.0,FALSE,Convexity of survival path,Modeling assumption
iq_per_bll,0.2,0.35,0.6,FALSE,IQ points gained per 1 ug/dL reduction in developmentally averaged child BLL,Canfield Lanphear Jusko Schnaas plus LMIC contextualization
earn_per_iq,0.003,0.005,0.01,FALSE,Percent earnings gain per IQ point,LMIC-centered range using Ozawa 2022 Grosse et al. 2002 and Hanushek-Woessmann 2008
p_att_anc,0.55,0.82,0.95,FALSE,Attendance at ANC contact,UNICEF ANC coverage benchmarks
p_att_imm,0.5,0.8,0.95,FALSE,Attendance at immunization contact,UNICEF or WHO immunization coverage benchmarks
p_att_birth,0.4,0.65,0.85,FALSE,Institutional delivery rate,UNICEF institutional delivery benchmarks
p_att_infant_kohl_contact,0.55,0.8,0.95,FALSE,Infant kohl contact via either facility delivery or early immunization catch-up,Union of correlated infant contacts based on UNICEF delivery and immunization benchmarks
fetal_transfer_coeff,0.7,0.8,0.95,FALSE,Fraction of maternal lead reduction transmitted to fetal lead reduction,Inferred from maternal-cord lead relationship
lactational_transfer_coeff,0.1,0.2,0.35,FALSE,Fraction of maternal lead reduction transmitted to the breastfed infant in the milk-dominated first year,Structural bridge from maternal blood lead to early-infancy exposure
share_prenatal_last4m,0.2,0.3,0.4,FALSE,Share of developmental cognitive harm attributed to the last four months in utero,Timing-weight assumption informed by prenatal lead literature
share_year1,0.2,0.3,0.4,FALSE,Share of developmental cognitive harm attributed to the first year of life,Timing-weight assumption informed by infancy lead literature
prev_pot,0.5,0.75,0.9,FALSE,Unsafe pot prevalence among target households,Hotspot prevalence assumption with improved sentinel targeting
prev_kohl_maternal,0.4,0.7,0.85,FALSE,Unsafe maternal kohl prevalence,Synthesized from survey and ethnographic evidence with hotspot targeting
prev_kohl_infant,0.4,0.7,0.85,FALSE,Unsafe infant kohl prevalence among infants,Synthesized from survey and ethnographic evidence with hotspot targeting
prev_utensils,0.4,0.6,0.8,FALSE,Unsafe utensil prevalence among target households,Context-specific prevalence assumption with improved sentinel targeting
fidelity,0.3,0.6,0.8,FALSE,Delivery fidelity,Implementation assumption with service-delivery support
adherence,0.3,0.5,0.6,FALSE,Complete substitution or adherence,Implementation assumption
p_voucher_issued,0.4,0.7,0.9,FALSE,Voucher successfully issued at ANC,Program implementation assumption
p_merchant_stock,0.6,0.8,0.95,FALSE,Approved merchant with voucher accepts and has pot in stock,Program implementation assumption
p_redeem,0.6,0.75,0.9,FALSE,Voucher redeemed at merchant,Voucher redemption literature and high-value durable good assumption
p_use_targeted_after_redemption,0.75,0.9,0.98,FALSE,Redeemed pot used by pregnant woman or toddler,Conditional use after self-targeted redemption
residual_unsafe_use_harm,0.05,0.15,0.35,FALSE,Share of pot benefit lost from continued unsafe pot use for target users,Product stacking penalty
labor_entry_age,14,18,20,TRUE,Age labor earnings begin,Economic timing assumption
mom_age,18,25,35,FALSE,Age of mother at intervention,DHS-style fertility timing assumption
gp_age,42,52,70,FALSE,Age of grandparent at intervention,Demographic assumption
gp_coresidence_prob,0.2,0.5,0.7,FALSE,Probability grandparent coresides,DHS or household survey synthesis
bll_pot_child,1,4,10,FALSE,Child BLL reduction from pot if fully effective during active use,Widened synthesis assumption from cookware exposure evidence
bll_pot_mother,0.5,2,5,FALSE,Mother BLL reduction from pot if fully effective during active use,Widened synthesis assumption from cookware exposure evidence
bll_mug_child,0.5,2,6,FALSE,Child BLL reduction from mug or spoon if fully effective during active use,Informed extrapolation with widened uncertainty
bll_calc_mother,0.2,0.4,1.2,FALSE,Mother BLL reduction from calcium if fully effective,Ettinger et al. 2009
bll_maternal_kohl_mother,0.5,2,5,FALSE,Mother BLL reduction from safe maternal kohl if fully effective,Widened synthesized kohl evidence
bll_infant_kohl_child,1.5,5,10,FALSE,Child BLL reduction from safe infant kohl if fully effective during active use,Widened synthesized kohl evidence
cost_pot_direct,8,12,15,TRUE,Direct product cost of lead-safe pot,Program design costing
cost_utensils_direct,1,1.5,2.5,TRUE,Direct product cost of safe utensils,Program design costing
cost_calcium_direct,1.25,1.75,3.0,TRUE,Direct product cost of calcium course,Program design costing
cost_kohl_maternal_direct,0.4,0.75,1.25,TRUE,Direct product cost of maternal kohl substitute,Program design costing
cost_kohl_infant_direct,1.2,1.5,3.0,TRUE,Direct product cost of infant kohl substitute,Program design costing
cost_explain,0.3,0.35,0.5,TRUE,Counseling cost per eligible contact,CHW time assumption
cost_bulk_distribution_margin,3.0,4.0,5.0,TRUE,Bulk distribution cost plus retailer margin for pot voucher redemption,Program logistics and merchant margin assumption
cost_dist_light,0.25,0.5,1.0,TRUE,Distribution cost for light items,Program logistics assumption
logistics_markup,0.3,0.6,1.0,TRUE,Program markup covering wastage supply-chain overhead supervision and health-system administration,Program logistics and administration assumption
district_fixed_cost,20000,50000,100000,TRUE,District fixed cost,Program design assumption
district_births,4000,20000,50000,FALSE,Births per district,Administrative planning assumption
preeclampsia_base_prev,0.02,0.04,0.06,FALSE,Baseline preeclampsia prevalence,WHO-compatible global burden range
preeclampsia_or,1.016,1.018,1.02,FALSE,Odds ratio per 1 ug/dL lead increase,Poropat et al. 2018
preterm_base_prev,0.08,0.12,0.15,FALSE,Baseline preterm prevalence,WHO preterm birth fact sheet
preterm_or,1.02,1.025,1.03,FALSE,Odds ratio per 1 ug/dL lead increase,Prenatal lead birth-outcome synthesis
preeclampsia_timing_mult,0.5,0.75,1.0,FALSE,Share of full maternal BLL reduction relevant to preeclampsia once intervention starts mid-pregnancy,Timing adjustment for later pregnancy risk window
preterm_timing_mult,0.4,0.7,1.0,FALSE,Share of full maternal BLL reduction relevant to preterm birth once intervention starts mid-pregnancy,Timing adjustment for pregnancy-window uncertainty
neonatal_daly_mult,0.5,0.75,1.0,FALSE,DALYs per preterm case averted,Reduced-form valuation assumption
cvd_daly_per_ug_lifetime,0.002,0.0033,0.005,FALSE,Lifetime adult DALYs per 1 ug/dL sustained reduction,Lead-CVD burden modeling assumption
vsly_mult,1.0,1.6,3.0,FALSE,VSLY multiplier relative to annual productivity,Robinson et al. style valuation range
preeclampsia_daly_wt,0.001,0.003,0.005,FALSE,DALYs per preeclampsia case,Reduced-form valuation assumption
years_pot,2,4,5,FALSE,Years pot protects,Program duration assumption with widened uncertainty
years_kohl_maternal,0.5,1,2,FALSE,Years maternal kohl protects,Program duration assumption with widened uncertainty
years_kohl_infant,0.5,1,2,FALSE,Years infant kohl protects,Program duration assumption with widened uncertainty
years_utensils,1,2,3,FALSE,Years utensils protect,Program duration assumption with widened uncertainty
years_calcium,0.25,0.5,1,FALSE,Years calcium protects,Program duration assumption with widened uncertainty
adult_reference_years,65,65,65,FALSE,Reference years for adult cumulative exposure,Modeling assumption
gp_pot_effect_multiplier,0.5,0.5,0.5,FALSE,Grandparent gets half the pot effect,Modeling assumption
"""


PARAMS = pd.read_csv(io.StringIO(CSV_DATA)).set_index("variable")

PROBABILITY_PARAMS = [
    "labor_share_income",
    "discount",
    "p18",
    "p65",
    "earn_per_iq",
    "p_att_anc",
    "p_att_imm",
    "p_att_birth",
    "p_att_infant_kohl_contact",
    "fetal_transfer_coeff",
    "lactational_transfer_coeff",
    "share_prenatal_last4m",
    "share_year1",
    "prev_pot",
    "prev_kohl_maternal",
    "prev_kohl_infant",
    "prev_utensils",
    "fidelity",
    "adherence",
    "p_voucher_issued",
    "p_merchant_stock",
    "p_redeem",
    "p_use_targeted_after_redemption",
    "residual_unsafe_use_harm",
    "gp_coresidence_prob",
    "preeclampsia_base_prev",
    "preeclampsia_timing_mult",
    "preterm_base_prev",
    "preterm_timing_mult",
]


def draw_pert(name, rng, size=N_TRIALS):
    row = PARAMS.loc[name]
    min_v = float(row["min"])
    mode_v = float(row["mode"])
    max_v = float(row["max"])

    width = max_v - min_v
    if width == 0:
        return np.full(size, mode_v)

    a = 1 + 4 * (mode_v - min_v) / width
    b = 1 + 4 * (max_v - mode_v) / width
    return stats.beta.rvs(a, b, size=size, random_state=rng) * width + min_v


def induce_rank_correlation(x, y, corr_val, rng):
    size = len(x)
    z = rng.normal(size=(2, size))
    L = cholesky([[1.0, corr_val], [corr_val, 1.0]], lower=True)
    z_corr = L @ z
    ranks = np.argsort(np.argsort(z_corr, axis=1), axis=1)
    return np.sort(x)[ranks[0]], np.sort(y)[ranks[1]]


def induce_shared_rank_factor(draw_dict, names, corr_val, rng):
    """Impose positive dependence across several implementation steps."""
    size = len(draw_dict[names[0]])
    shared = rng.normal(size=size)
    for name in names:
        indiv = rng.normal(size=size)
        mixed = np.sqrt(corr_val) * shared + np.sqrt(1 - corr_val) * indiv
        ranks = np.argsort(np.argsort(mixed))
        draw_dict[name] = np.sort(draw_dict[name])[ranks]
    return draw_dict


def odds_to_risk(base_prev, odds_ratio, delta_bll):
    base_odds = base_prev / (1 - base_prev)
    new_odds = base_odds / np.power(odds_ratio, delta_bll)
    return new_odds / (1 + new_odds)


def calculate_pv_earnings(productivity, growth, discount, p18, p65, k_shape, labor_entry_age):
    x = (1 + growth) / (1 + discount)
    t_work = np.round(65 - labor_entry_age + 1).astype(int)
    t_idx = np.arange(1, 53).reshape(-1, 1)
    st = p18 - (p18 - p65) * (np.maximum(0, t_idx - 1) / 47.0) ** k_shape
    pv_sum = np.sum((t_idx <= t_work) * st * x ** (t_idx - 1), axis=0)
    return productivity * (x ** labor_entry_age) * pv_sum


def draw_params(seed=RNG_SEED):
    rng = np.random.default_rng(seed)
    draws = {name: draw_pert(name, rng) for name in PARAMS.index}
    # Pot implementation steps are likely to rise and fall together in the same
    # local delivery environment, so we add a modest positive dependence across
    # attendance, fidelity, issuance, stock, redemption, and targeted use.
    draws = induce_shared_rank_factor(
        draws,
        [
            "p_att_anc",
            "fidelity",
            "adherence",
            "p_voucher_issued",
            "p_merchant_stock",
            "p_redeem",
            "p_use_targeted_after_redemption",
        ],
        0.30,
        rng,
    )
    draws["fidelity"], draws["adherence"] = induce_rank_correlation(
        draws["fidelity"], draws["adherence"], 0.45, rng
    )
    draws = induce_shared_rank_factor(
        draws,
        ["p_att_infant_kohl_contact", "fidelity", "adherence"],
        0.20,
        rng,
    )
    draws["share_prenatal_last4m"], draws["share_year1"], draws["share_years2to5"] = draw_dev_shares(
        draws["share_prenatal_last4m"],
        draws["share_year1"],
        rng,
    )
    draws["is_daughter"] = rng.binomial(1, 0.5, N_TRIALS)
    validate_draws(draws)
    return draws


def draw_dev_shares(prenatal_draw, year1_draw, rng):
    """Draw developmental harm shares that sum to one.

    The third share, covering ages 2 to 5, is defined residually so the
    developmental weights always sum to one. We resample the first two shares
    until the residual lies in the target range [0.30, 0.60].
    """

    prenatal = np.array(prenatal_draw, copy=True)
    year1 = np.array(year1_draw, copy=True)
    years2to5 = 1.0 - prenatal - year1

    mask = (years2to5 < 0.30) | (years2to5 > 0.60)
    while np.any(mask):
        prenatal[mask] = draw_pert("share_prenatal_last4m", rng, size=int(mask.sum()))
        year1[mask] = draw_pert("share_year1", rng, size=int(mask.sum()))
        years2to5[mask] = 1.0 - prenatal[mask] - year1[mask]
        mask = (years2to5 < 0.30) | (years2to5 > 0.60)

    return prenatal, year1, years2to5


def validate_draws(p):
    for name in PROBABILITY_PARAMS:
        vals = np.asarray(p[name])
        if np.any((vals < 0) | (vals > 1)):
            raise ValueError(f"{name} has values outside [0,1]")

    total_share = p["share_prenatal_last4m"] + p["share_year1"] + p["share_years2to5"]
    if not np.allclose(total_share, 1.0, atol=1e-8):
        raise ValueError("Developmental shares do not sum to one")

    if np.any((p["share_years2to5"] < 0.30) | (p["share_years2to5"] > 0.60)):
        raise ValueError("Residual developmental share is outside [0.30, 0.60]")


def validate_results(df, package_name):
    if (df["cost"] <= 0).any():
        raise ValueError(f"{package_name}: nonpositive costs detected")

    identity_mn = np.abs(df["bcr_plus_mn"] - (df["benefit_earnings"] + df["benefit_maternal_neonatal"]) / df["cost"])
    identity_total = np.abs(
        df["bcr_total"]
        - (df["benefit_earnings"] + df["benefit_maternal_neonatal"] + df["benefit_cvd"]) / df["cost"]
    )
    if float(identity_mn.max()) > 1e-10:
        raise ValueError(f"{package_name}: BCR plus maternal/neonatal identity check failed")
    if float(identity_total.max()) > 1e-10:
        raise ValueError(f"{package_name}: total BCR identity check failed")

    for column in [
        "cost",
        "benefit_earnings",
        "benefit_maternal_neonatal",
        "benefit_cvd",
        "delta_iq",
    ]:
        if not np.isfinite(df[column]).all():
            raise ValueError(f"{package_name}: non-finite values detected in {column}")


def build_validation_report(package_name, df):
    return {
        "package": package_name,
        "n_trials": int(len(df)),
        "median_cost": float(df["cost"].median()),
        "share_cost_positive": float((df["cost"] > 0).mean()),
        "max_bcr_plus_mn_identity_error": float(
            np.abs(df["bcr_plus_mn"] - (df["benefit_earnings"] + df["benefit_maternal_neonatal"]) / df["cost"]).max()
        ),
        "max_bcr_total_identity_error": float(
            np.abs(
                df["bcr_total"]
                - (df["benefit_earnings"] + df["benefit_maternal_neonatal"] + df["benefit_cvd"]) / df["cost"]
            ).max()
        ),
        "share_dev_shares_sum_to_one": float(
            np.isclose(
                df["share_prenatal_last4m"] + df["share_year1"] + df["share_years2to5"],
                1.0,
                atol=1e-8,
            ).mean()
        ),
    }


def mode_params():
    params = {name: np.array([float(row["mode"])]) for name, row in PARAMS.iterrows()}
    params["share_years2to5"] = 1.0 - params["share_prenatal_last4m"] - params["share_year1"]
    params["is_daughter"] = np.array([0.5])
    validate_draws(params)
    return params


def add_common_economic_fields(p):
    # Productivity is anchored to the labour share of GDP, not formal labour-force
    # participation, so informal and self-employed market production are included.
    productivity = p["gdp_ppp_per_capita"] * p["labor_share_income"]
    pv_earnings = calculate_pv_earnings(
        productivity,
        p["growth"],
        p["discount"],
        p["p18"],
        p["p65"],
        p["k"],
        p["labor_entry_age"],
    )
    fixed_cost_per_birth = p["district_fixed_cost"] / p["district_births"]
    vsly = p["vsly_mult"] * productivity
    return productivity, pv_earnings, fixed_cost_per_birth, vsly


def postnatal_cog_weight(duration_years, share_year1, share_years2to5, start_offset_years=0.0):
    """Map a postnatal pathway duration into developmental-harm weights.

    Assumption:
    - Year 1 receives its own developmental share.
    - Years 2 to 5 share the remaining postnatal developmental weight evenly.
    """

    effective_duration = np.clip(duration_years, 0, None)
    year1_window = np.clip(1.0 - start_offset_years, 0, 1)
    covered_year1 = np.clip(effective_duration, 0, year1_window)
    remaining_duration = np.clip(effective_duration - covered_year1, 0, None)
    covered_years2to5 = np.clip(remaining_duration, 0, 4)
    return share_year1 * covered_year1 + share_years2to5 * (covered_years2to5 / 4.0)


def simulate_kitchen(p):
    productivity, pv_earnings, fixed_cost_per_birth, vsly = add_common_economic_fields(p)

    p_pot = (
        p["prev_pot"]
        * p["p_att_anc"]
        * p["p_voucher_issued"]
        * p["p_merchant_stock"]
        * p["p_redeem"]
        * p["p_use_targeted_after_redemption"]
    )
    pot_penalty = 1 - p["residual_unsafe_use_harm"]
    # Interpretation 1: the child pot effect is postnatal-only. Any prenatal
    # child benefit from lower maternal lead is modeled separately below.
    eff_bll_pot_child = p["bll_pot_child"] * p_pot * pot_penalty
    eff_bll_pot_mother = p["bll_pot_mother"] * p_pot * pot_penalty

    p_calcium = p["p_att_anc"] * p["fidelity"] * p["adherence"]
    eff_bll_calc_mother = p["bll_calc_mother"] * p_calcium

    p_utensils = p["prev_utensils"] * p["p_att_imm"] * p["fidelity"] * p["adherence"]
    eff_bll_utensils_child = p["bll_mug_child"] * p_utensils

    prenatal_child_bll = p["fetal_transfer_coeff"] * (eff_bll_pot_mother + eff_bll_calc_mother)
    lactational_child_bll = p["lactational_transfer_coeff"] * eff_bll_pot_mother
    total_child_bll = eff_bll_pot_child + eff_bll_utensils_child + prenatal_child_bll + lactational_child_bll
    child_bll_cog_equiv = (
        prenatal_child_bll * p["share_prenatal_last4m"]
        + lactational_child_bll * p["share_year1"] * np.clip(p["years_pot"], 0, 1)
        + eff_bll_pot_child
        * postnatal_cog_weight(p["years_pot"], p["share_year1"], p["share_years2to5"], start_offset_years=0.5)
        + eff_bll_utensils_child
        * postnatal_cog_weight(p["years_utensils"], p["share_year1"], p["share_years2to5"], start_offset_years=0.5)
    )
    delta_iq = child_bll_cog_equiv * p["iq_per_bll"]
    ben_earnings = pv_earnings * p["earn_per_iq"] * delta_iq
    cost = np.array(fixed_cost_per_birth, copy=True)
    cost += p["cost_explain"] * p["p_att_anc"]
    cost += (
        (p["cost_pot_direct"] + p["cost_bulk_distribution_margin"])
        * (1 + p["logistics_markup"])
        * p["p_att_anc"]
        * p["p_voucher_issued"]
        * p["p_merchant_stock"]
        * p["p_redeem"]
    )
    cost += p["cost_explain"] * p["p_att_anc"]
    cost += p["cost_dist_light"] * (1 + p["logistics_markup"]) * p["p_att_anc"] * p["fidelity"]
    cost += p["cost_calcium_direct"] * (1 + p["logistics_markup"]) * p["p_att_anc"] * p["fidelity"]
    cost += p["cost_explain"] * p["p_att_imm"]
    cost += p["cost_dist_light"] * (1 + p["logistics_markup"]) * p["p_att_imm"] * p["fidelity"]
    cost += p["cost_utensils_direct"] * (1 + p["logistics_markup"]) * p["p_att_imm"] * p["fidelity"]

    maternal_bll_for_preeclampsia = (eff_bll_pot_mother + eff_bll_calc_mother) * p["preeclampsia_timing_mult"]
    preeclampsia_new_prev = odds_to_risk(
        p["preeclampsia_base_prev"], p["preeclampsia_or"], maternal_bll_for_preeclampsia
    )
    delta_preeclampsia = p["preeclampsia_base_prev"] - preeclampsia_new_prev
    ben_preeclampsia = delta_preeclampsia * p["preeclampsia_daly_wt"] * vsly

    maternal_bll_for_preterm = (eff_bll_pot_mother + eff_bll_calc_mother) * p["preterm_timing_mult"]
    preterm_new_prev = odds_to_risk(
        p["preterm_base_prev"], p["preterm_or"], maternal_bll_for_preterm
    )
    delta_preterm = p["preterm_base_prev"] - preterm_new_prev
    ben_preterm = delta_preterm * p["neonatal_daly_mult"] * vsly
    ben_maternal_neonatal = ben_preeclampsia + ben_preterm

    ben_cvd_gp = (
        eff_bll_pot_child
        * p["cvd_daly_per_ug_lifetime"]
        * p["years_pot"]
        * vsly
        * p["gp_pot_effect_multiplier"]
        * np.power(1 + p["discount"], -np.maximum(0, p["adult_reference_years"] - p["gp_age"]))
        * p["gp_coresidence_prob"]
    )
    ben_cvd_mom = (
        (eff_bll_pot_mother * p["years_pot"] + eff_bll_calc_mother * p["years_calcium"])
        * p["cvd_daly_per_ug_lifetime"]
        * vsly
        * np.power(1 + p["discount"], -np.maximum(0, p["adult_reference_years"] - p["mom_age"]))
    )
    ben_cvd = ben_cvd_mom + ben_cvd_gp

    return pd.DataFrame(
        {
            "package": "kitchen",
            "cost": cost,
            "pv_earnings": pv_earnings,
            "benefit_earnings": ben_earnings,
            "benefit_maternal_neonatal": ben_maternal_neonatal,
            "benefit_cvd": ben_cvd,
            "bcr_cog": ben_earnings / cost,
            "bcr_plus_mn": (ben_earnings + ben_maternal_neonatal) / cost,
            "bcr_total": (ben_earnings + ben_maternal_neonatal + ben_cvd) / cost,
            "eff_bll_pot_child": eff_bll_pot_child,
            "eff_bll_pot_mother": eff_bll_pot_mother,
            "eff_bll_calcium_mother": eff_bll_calc_mother,
            "eff_bll_utensils_child": eff_bll_utensils_child,
            "eff_bll_prenatal_child": prenatal_child_bll,
            "eff_bll_lactational_child": lactational_child_bll,
            "eff_bll_child_total": total_child_bll,
            "eff_bll_child_cog_equiv": child_bll_cog_equiv,
            "delta_iq": delta_iq,
            "delta_preeclampsia_pp": delta_preeclampsia,
            "delta_preterm_pp": delta_preterm,
            "maternal_bll_for_preeclampsia": maternal_bll_for_preeclampsia,
            "maternal_bll_for_preterm": maternal_bll_for_preterm,
            "p_success_pot": p_pot,
            "p_success_calcium": p_calcium,
            "p_success_utensils": p_utensils,
            "share_prenatal_last4m": p["share_prenatal_last4m"],
            "share_year1": p["share_year1"],
            "share_years2to5": p["share_years2to5"],
        }
    )


def simulate_kohl(p, infant_kohl_multiplier=None):
    productivity, pv_earnings, fixed_cost_per_birth, vsly = add_common_economic_fields(p)
    if infant_kohl_multiplier is None:
        infant_kohl_multiplier = np.ones_like(p["prev_kohl_infant"])

    p_maternal = p["prev_kohl_maternal"] * p["p_att_anc"] * p["fidelity"] * p["adherence"]
    p_infant = (
        p["prev_kohl_infant"]
        * p["p_att_infant_kohl_contact"]
        * p["fidelity"]
        * p["adherence"]
    )

    eff_bll_kohl_mother = p["bll_maternal_kohl_mother"] * p_maternal
    eff_bll_kohl_infant = infant_kohl_multiplier * p["bll_infant_kohl_child"] * p_infant
    prenatal_child_bll = p["fetal_transfer_coeff"] * eff_bll_kohl_mother
    lactational_child_bll = p["lactational_transfer_coeff"] * eff_bll_kohl_mother
    total_child_bll = prenatal_child_bll + lactational_child_bll + eff_bll_kohl_infant
    child_bll_cog_equiv = (
        prenatal_child_bll * p["share_prenatal_last4m"]
        + lactational_child_bll * p["share_year1"] * np.clip(p["years_kohl_maternal"], 0, 1)
        + eff_bll_kohl_infant
        * postnatal_cog_weight(p["years_kohl_infant"], p["share_year1"], p["share_years2to5"])
    )

    delta_iq = child_bll_cog_equiv * p["iq_per_bll"]
    ben_earnings = pv_earnings * p["earn_per_iq"] * delta_iq
    cost = np.array(fixed_cost_per_birth, copy=True)
    cost += p["cost_explain"] * p["p_att_anc"]
    cost += p["cost_dist_light"] * (1 + p["logistics_markup"]) * p["p_att_anc"] * p["fidelity"]
    cost += (
        p["cost_kohl_maternal_direct"]
        * np.clip(p["years_kohl_maternal"], 0, None)
        * (1 + p["logistics_markup"])
        * p["p_att_anc"]
        * p["fidelity"]
    )
    cost += p["cost_explain"] * p["p_att_infant_kohl_contact"]
    cost += (
        p["cost_dist_light"]
        * (1 + p["logistics_markup"])
        * p["p_att_infant_kohl_contact"]
        * p["fidelity"]
    )
    cost += (
        infant_kohl_multiplier
        * p["cost_kohl_infant_direct"]
        * (1 + p["logistics_markup"])
        * p["p_att_infant_kohl_contact"]
        * p["fidelity"]
    )

    maternal_bll_for_preeclampsia = eff_bll_kohl_mother * p["preeclampsia_timing_mult"]
    preeclampsia_new_prev = odds_to_risk(
        p["preeclampsia_base_prev"], p["preeclampsia_or"], maternal_bll_for_preeclampsia
    )
    delta_preeclampsia = p["preeclampsia_base_prev"] - preeclampsia_new_prev
    ben_preeclampsia = delta_preeclampsia * p["preeclampsia_daly_wt"] * vsly

    maternal_bll_for_preterm = eff_bll_kohl_mother * p["preterm_timing_mult"]
    preterm_new_prev = odds_to_risk(p["preterm_base_prev"], p["preterm_or"], maternal_bll_for_preterm)
    delta_preterm = p["preterm_base_prev"] - preterm_new_prev
    ben_preterm = delta_preterm * p["neonatal_daly_mult"] * vsly
    ben_maternal_neonatal = ben_preeclampsia + ben_preterm

    ben_cvd_mom = (
        eff_bll_kohl_mother
        * p["cvd_daly_per_ug_lifetime"]
        * p["years_kohl_maternal"]
        * vsly
        * np.power(1 + p["discount"], -np.maximum(0, p["adult_reference_years"] - p["mom_age"]))
    )

    return pd.DataFrame(
        {
            "package": "kohl",
            "cost": cost,
            "pv_earnings": pv_earnings,
            "benefit_earnings": ben_earnings,
            "benefit_maternal_neonatal": ben_maternal_neonatal,
            "benefit_cvd": ben_cvd_mom,
            "bcr_cog": ben_earnings / cost,
            "bcr_plus_mn": (ben_earnings + ben_maternal_neonatal) / cost,
            "bcr_total": (ben_earnings + ben_maternal_neonatal + ben_cvd_mom) / cost,
            "eff_bll_kohl_mother": eff_bll_kohl_mother,
            "eff_bll_kohl_infant": eff_bll_kohl_infant,
            "eff_bll_prenatal_child": prenatal_child_bll,
            "eff_bll_lactational_child": lactational_child_bll,
            "eff_bll_child_total": total_child_bll,
            "eff_bll_child_cog_equiv": child_bll_cog_equiv,
            "delta_iq": delta_iq,
            "delta_preeclampsia_pp": delta_preeclampsia,
            "delta_preterm_pp": delta_preterm,
            "maternal_bll_for_preeclampsia": maternal_bll_for_preeclampsia,
            "maternal_bll_for_preterm": maternal_bll_for_preterm,
            "p_success_kohl_maternal": p_maternal,
            "p_success_kohl_infant": p_infant,
            "infant_kohl_multiplier": infant_kohl_multiplier,
            "share_prenatal_last4m": p["share_prenatal_last4m"],
            "share_year1": p["share_year1"],
            "share_years2to5": p["share_years2to5"],
        }
    )


def summarize_results(df, bcr_col):
    series = df[bcr_col]
    return {
        "median": float(series.median()),
        "p5": float(series.quantile(0.05)),
        "p95": float(series.quantile(0.95)),
        "fail_pct": float((series < 1).mean() * 100),
    }


def summarize_case(name, df):
    cog = summarize_results(df, "bcr_cog")
    total = summarize_results(df, "bcr_total")
    return {
        "package_case": name,
        "bcr_cog_median": cog["median"],
        "bcr_cog_p5": cog["p5"],
        "bcr_cog_p95": cog["p95"],
        "bcr_cog_fail_pct": cog["fail_pct"],
        "bcr_total_median": total["median"],
        "bcr_total_p5": total["p5"],
        "bcr_total_p95": total["p95"],
        "bcr_total_fail_pct": total["fail_pct"],
    }


def scale_param_block(param_table, names, scale):
    updated = param_table.copy(deep=True)
    for name in names:
        for col in ["min", "mode", "max"]:
            updated.loc[name, col] = float(updated.loc[name, col]) * scale
    return updated


def export_disadvantaged_scenario(output_dir):
    scenario_params = [
        "gdp_ppp_per_capita",
        "growth",
        "p_att_anc",
        "fidelity",
        "p_merchant_stock",
        "p_redeem",
        "adherence",
    ]
    scale = 0.7
    baseline_rows = [
        summarize_case("Kitchen Baseline", simulate_kitchen(draw_params(seed=RNG_SEED))),
        summarize_case("Kohl Baseline", simulate_kohl(draw_params(seed=RNG_SEED + 2))),
    ]
    original_params = PARAMS.copy(deep=True)
    try:
        globals()["PARAMS"] = scale_param_block(PARAMS, scenario_params, scale)
        scenario_rows = [
            summarize_case("Kitchen Disadvantaged Setting", simulate_kitchen(draw_params(seed=RNG_SEED))),
            summarize_case("Kohl Disadvantaged Setting", simulate_kohl(draw_params(seed=RNG_SEED + 2))),
        ]
    finally:
        globals()["PARAMS"] = original_params
    baseline = pd.DataFrame(baseline_rows).set_index("package_case")
    scenario = pd.DataFrame(scenario_rows).set_index("package_case")
    comparison = pd.DataFrame(
        [
            {
                "package": "Kitchen",
                "baseline_bcr_cog_median": baseline.loc["Kitchen Baseline", "bcr_cog_median"],
                "scenario_bcr_cog_median": scenario.loc["Kitchen Disadvantaged Setting", "bcr_cog_median"],
                "pct_change_bcr_cog_median": (scenario.loc["Kitchen Disadvantaged Setting", "bcr_cog_median"] / baseline.loc["Kitchen Baseline", "bcr_cog_median"] - 1) * 100,
                "baseline_bcr_total_median": baseline.loc["Kitchen Baseline", "bcr_total_median"],
                "scenario_bcr_total_median": scenario.loc["Kitchen Disadvantaged Setting", "bcr_total_median"],
                "pct_change_bcr_total_median": (scenario.loc["Kitchen Disadvantaged Setting", "bcr_total_median"] / baseline.loc["Kitchen Baseline", "bcr_total_median"] - 1) * 100,
                "baseline_fail_pct_total": baseline.loc["Kitchen Baseline", "bcr_total_fail_pct"],
                "scenario_fail_pct_total": scenario.loc["Kitchen Disadvantaged Setting", "bcr_total_fail_pct"],
            },
            {
                "package": "Kohl",
                "baseline_bcr_cog_median": baseline.loc["Kohl Baseline", "bcr_cog_median"],
                "scenario_bcr_cog_median": scenario.loc["Kohl Disadvantaged Setting", "bcr_cog_median"],
                "pct_change_bcr_cog_median": (scenario.loc["Kohl Disadvantaged Setting", "bcr_cog_median"] / baseline.loc["Kohl Baseline", "bcr_cog_median"] - 1) * 100,
                "baseline_bcr_total_median": baseline.loc["Kohl Baseline", "bcr_total_median"],
                "scenario_bcr_total_median": scenario.loc["Kohl Disadvantaged Setting", "bcr_total_median"],
                "pct_change_bcr_total_median": (scenario.loc["Kohl Disadvantaged Setting", "bcr_total_median"] / baseline.loc["Kohl Baseline", "bcr_total_median"] - 1) * 100,
                "baseline_fail_pct_total": baseline.loc["Kohl Baseline", "bcr_total_fail_pct"],
                "scenario_fail_pct_total": scenario.loc["Kohl Disadvantaged Setting", "bcr_total_fail_pct"],
            },
        ]
    )
    parameter_compare = (
        original_params.loc[scenario_params, ["min", "mode", "max"]]
        .rename(columns=lambda c: f"baseline_{c}")
        .join(
            scale_param_block(original_params, scenario_params, scale)
            .loc[scenario_params, ["min", "mode", "max"]]
            .rename(columns={"min": "scenario_min", "mode": "scenario_mode", "max": "scenario_max"})
        )
    )
    parameter_compare.insert(0, "scaled_by", scale)
    pd.DataFrame(baseline_rows + scenario_rows).to_csv(output_dir / "summary_table_with_disadvantaged_scenario.csv", index=False)
    comparison.to_csv(output_dir / "disadvantaged_scenario_comparison.csv", index=False)
    parameter_compare.to_csv(output_dir / "disadvantaged_scenario_parameters.csv")


def export_half_effect_scenario(output_dir):
    scenario_params = [
        "bll_pot_child",
        "bll_pot_mother",
        "bll_mug_child",
        "bll_maternal_kohl_mother",
        "bll_infant_kohl_child",
    ]
    scale = 0.5
    baseline_rows = [
        summarize_case("Kitchen Baseline", simulate_kitchen(draw_params(seed=RNG_SEED))),
        summarize_case("Kohl Baseline", simulate_kohl(draw_params(seed=RNG_SEED + 2))),
    ]
    original_params = PARAMS.copy(deep=True)
    try:
        globals()["PARAMS"] = scale_param_block(PARAMS, scenario_params, scale)
        scenario_rows = [
            summarize_case("Kitchen Half Pot-Utensil Effect", simulate_kitchen(draw_params(seed=RNG_SEED))),
            summarize_case("Kohl Half Kohl Effect", simulate_kohl(draw_params(seed=RNG_SEED + 2))),
        ]
    finally:
        globals()["PARAMS"] = original_params
    baseline = pd.DataFrame(baseline_rows).set_index("package_case")
    scenario = pd.DataFrame(scenario_rows).set_index("package_case")
    comparison = pd.DataFrame(
        [
            {
                "package": "Kitchen",
                "baseline_bcr_cog_median": baseline.loc["Kitchen Baseline", "bcr_cog_median"],
                "scenario_bcr_cog_median": scenario.loc["Kitchen Half Pot-Utensil Effect", "bcr_cog_median"],
                "pct_change_bcr_cog_median": (scenario.loc["Kitchen Half Pot-Utensil Effect", "bcr_cog_median"] / baseline.loc["Kitchen Baseline", "bcr_cog_median"] - 1) * 100,
                "baseline_bcr_total_median": baseline.loc["Kitchen Baseline", "bcr_total_median"],
                "scenario_bcr_total_median": scenario.loc["Kitchen Half Pot-Utensil Effect", "bcr_total_median"],
                "pct_change_bcr_total_median": (scenario.loc["Kitchen Half Pot-Utensil Effect", "bcr_total_median"] / baseline.loc["Kitchen Baseline", "bcr_total_median"] - 1) * 100,
                "baseline_fail_pct_total": baseline.loc["Kitchen Baseline", "bcr_total_fail_pct"],
                "scenario_fail_pct_total": scenario.loc["Kitchen Half Pot-Utensil Effect", "bcr_total_fail_pct"],
            },
            {
                "package": "Kohl",
                "baseline_bcr_cog_median": baseline.loc["Kohl Baseline", "bcr_cog_median"],
                "scenario_bcr_cog_median": scenario.loc["Kohl Half Kohl Effect", "bcr_cog_median"],
                "pct_change_bcr_cog_median": (scenario.loc["Kohl Half Kohl Effect", "bcr_cog_median"] / baseline.loc["Kohl Baseline", "bcr_cog_median"] - 1) * 100,
                "baseline_bcr_total_median": baseline.loc["Kohl Baseline", "bcr_total_median"],
                "scenario_bcr_total_median": scenario.loc["Kohl Half Kohl Effect", "bcr_total_median"],
                "pct_change_bcr_total_median": (scenario.loc["Kohl Half Kohl Effect", "bcr_total_median"] / baseline.loc["Kohl Baseline", "bcr_total_median"] - 1) * 100,
                "baseline_fail_pct_total": baseline.loc["Kohl Baseline", "bcr_total_fail_pct"],
                "scenario_fail_pct_total": scenario.loc["Kohl Half Kohl Effect", "bcr_total_fail_pct"],
            },
        ]
    )
    parameter_compare = (
        original_params.loc[scenario_params, ["min", "mode", "max"]]
        .rename(columns=lambda c: f"baseline_{c}")
        .join(
            scale_param_block(original_params, scenario_params, scale)
            .loc[scenario_params, ["min", "mode", "max"]]
            .rename(columns={"min": "scenario_min", "mode": "scenario_mode", "max": "scenario_max"})
        )
    )
    parameter_compare.insert(0, "scaled_by", scale)
    pd.DataFrame(baseline_rows + scenario_rows).to_csv(output_dir / "summary_table_with_half_effect_scenario.csv", index=False)
    comparison.to_csv(output_dir / "half_effect_scenario_comparison.csv", index=False)
    parameter_compare.to_csv(output_dir / "half_effect_scenario_parameters.csv")


def modal_policy_scale_outputs(df_mode, births=1_000_000):
    row = df_mode.iloc[0]
    total_benefit = float(row["benefit_earnings"] + row["benefit_maternal_neonatal"] + row["benefit_cvd"])
    total_cost = float(row["cost"])
    return {
        "modal_cost_per_1m_births_usd": total_cost * births,
        "modal_total_benefit_per_1m_births_usd": total_benefit * births,
        "modal_net_benefit_per_1m_births_usd": (total_benefit - total_cost) * births,
    }


def export_trials(df, output_dir, stem):
    csv_path = output_dir / f"{stem}_trials.csv.gz"
    df.to_csv(csv_path, index=False, compression="gzip")
    return csv_path


def plot_single_bcr(df, package_name, output_dir):
    if plt is None:
        return
    x_cap = min(float(df["bcr_cog"].quantile(0.99)), 70.0)
    plt.figure(figsize=(8, 5))
    plt.hist(df["bcr_cog"], bins=80, density=True, alpha=0.55, color="tab:blue")
    plt.axvline(1.0, linestyle="--", color="black", linewidth=1.2, label="Breakeven")
    plt.axvline(df["bcr_cog"].median(), linestyle="--", color="tab:red", linewidth=1.1, label="Median")
    plt.xlim(0, max(2, x_cap))
    plt.xlabel("Benefit-Cost Ratio")
    plt.ylabel("Density")
    plt.title(f"Probabilistic Sensitivity Analysis of the {package_name} Package")
    plt.legend()
    plt.tight_layout()
    plt.savefig(output_dir / f"{package_name.lower()}_bcr_baseline.png", dpi=220)
    plt.close()


def plot_bcr_comparison(kitchen_df, kohl_df, output_dir):
    if plt is None:
        return
    x_cap = min(float(max(kitchen_df["bcr_cog"].quantile(0.99), kohl_df["bcr_cog"].quantile(0.99))), 70.0)
    kitchen_median = float(kitchen_df["bcr_cog"].median())
    kohl_median = float(kohl_df["bcr_cog"].median())
    plt.figure(figsize=(8, 5))
    plt.hist(kitchen_df["bcr_cog"], bins=80, density=True, alpha=0.45, label="Kitchen", color="tab:blue")
    plt.hist(kohl_df["bcr_cog"], bins=80, density=True, alpha=0.45, label="Kohl", color="tab:orange")
    plt.axvline(1.0, linestyle="--", color="black", linewidth=1.2, label="Breakeven")
    plt.axvline(kitchen_median, linestyle="--", color="tab:blue", linewidth=1.4)
    plt.axvline(kohl_median, linestyle="--", color="tab:orange", linewidth=1.4)
    plt.xlim(0, max(2, x_cap))
    plt.xlabel("Benefit-Cost Ratio")
    plt.ylabel("Density")
    plt.title("Probabilistic Sensitivity Analysis of Sentinel Firewall Packages")
    ymax = plt.ylim()[1]
    kitchen_text_y = ymax * 0.92
    kohl_text_y = ymax * 0.80
    if abs(kitchen_median - kohl_median) < 0.6:
        kitchen_text_y = ymax * 0.92
        kohl_text_y = ymax * 0.70
    plt.text(
        kitchen_median,
        kitchen_text_y,
        f"Kitchen median = {kitchen_median:.2f}",
        color="tab:blue",
        rotation=90,
        va="top",
        ha="right",
        fontsize=9,
        bbox={"boxstyle": "round,pad=0.2", "facecolor": "white", "edgecolor": "tab:blue", "alpha": 0.85},
    )
    plt.text(
        kohl_median,
        kohl_text_y,
        f"Kohl median = {kohl_median:.2f}",
        color="tab:orange",
        rotation=90,
        va="top",
        ha="left",
        fontsize=9,
        bbox={"boxstyle": "round,pad=0.2", "facecolor": "white", "edgecolor": "tab:orange", "alpha": 0.85},
    )
    plt.legend()
    plt.tight_layout()
    plt.savefig(output_dir / "package_bcr_comparison.png", dpi=200)
    plt.close()


def draw_box(ax, xy, wh, text, facecolor="#5B84C4", edgecolor="#3D629B", fontsize=10, textcolor="white"):
    x, y = xy
    w, h = wh
    patch = FancyBboxPatch(
        (x, y),
        w,
        h,
        boxstyle="round,pad=0.02,rounding_size=0.03",
        linewidth=1.4,
        facecolor=facecolor,
        edgecolor=edgecolor,
    )
    ax.add_patch(patch)
    ax.text(x + w / 2, y + h / 2, text, ha="center", va="center", color=textcolor, fontsize=fontsize, wrap=True)
    return patch


def arrow_between(ax, start, end, color="#95A9C6", lw=2.2, ms=18):
    ax.add_patch(
        FancyArrowPatch(
            start,
            end,
            arrowstyle="-|>",
            mutation_scale=ms,
            linewidth=lw,
            color=color,
            shrinkA=0,
            shrinkB=0,
            connectionstyle="arc3",
        )
    )


def plot_main_pathway_figure(output_dir):
    if plt is None or FancyBboxPatch is None:
        return

    fig = plt.figure(figsize=(11, 8))
    gs = fig.add_gridspec(2, 1, height_ratios=[1, 1], hspace=0.2)
    ax1 = fig.add_subplot(gs[0, 0])
    ax2 = fig.add_subplot(gs[1, 0])
    for ax in (ax1, ax2):
        ax.set_xlim(0, 1)
        ax.set_ylim(0, 1)
        ax.axis("off")

    # Panel A: core causal pathway
    ax1.text(0.02, 0.95, "Figure 1A. Core causal pathway", fontsize=14, fontweight="bold", va="top")
    a = draw_box(ax1, (0.03, 0.48), (0.2, 0.23), "Safe-product\nsubstitution", facecolor="#4F7DB8")
    b = draw_box(
        ax1,
        (0.31, 0.63),
        (0.25, 0.18),
        "Lower maternal blood lead\nwhen pregnant and nursing",
        facecolor="#5E6CB2",
    )
    c = draw_box(ax1, (0.31, 0.25), (0.25, 0.18), "Lower child blood\nlead level", facecolor="#5E6CB2")
    d = draw_box(ax1, (0.64, 0.41), (0.19, 0.18), "Improved child\nneurodevelopment\nand higher IQ", facecolor="#4F7DB8")
    e = draw_box(ax1, (0.87, 0.41), (0.1, 0.18), "Higher\nlifetime\nearnings", facecolor="#B46A3C", edgecolor="#8D4F25")

    arrow_between(ax1, (0.23, 0.59), (0.31, 0.72))
    arrow_between(ax1, (0.23, 0.59), (0.31, 0.34))
    arrow_between(ax1, (0.435, 0.63), (0.435, 0.43), color="#7F92B0")
    arrow_between(ax1, (0.56, 0.50), (0.64, 0.50))
    arrow_between(ax1, (0.83, 0.50), (0.87, 0.50), color="#97B648")

    ax1.text(0.04, 0.18, "Calcium and maternal kohl -> maternal pathway", fontsize=8.8, color="#384860")
    ax1.text(0.04, 0.11, "Child utensils and infant kohl -> child pathway", fontsize=8.8, color="#384860")
    ax1.text(0.04, 0.04, "Safe pots -> both maternal and child pathways", fontsize=8.8, color="#384860")

    # Panel B: implementation chain
    ax2.text(0.02, 0.95, "Figure 1B. Implementation chain", fontsize=14, fontweight="bold", va="top")
    boxes = [
        ((0.02, 0.58), (0.16, 0.18), "Hazardous product\nis common in\ntargeted households"),
        ((0.21, 0.58), (0.18, 0.18), "Pregnant woman or\ninfant is reached\nthrough ANC, birth,\nor immunization"),
        ((0.42, 0.58), (0.16, 0.18), "Health system\ndelivers\nintervention"),
        ((0.61, 0.58), (0.16, 0.18), "Voucher or product\nis successfully\nobtained"),
        ((0.12, 0.19), (0.2, 0.18), "Caregiver uses\nsafe product for\npregnant woman or\nyoung child"),
        ((0.4, 0.19), (0.17, 0.18), "Residual unsafe\nuse is limited"),
        ((0.67, 0.19), (0.21, 0.18), "Realized population-level\nblood lead reduction"),
    ]
    for xy, wh, txt in boxes:
        draw_box(ax2, xy, wh, txt, facecolor="#4F7DB8")
    arrow_between(ax2, (0.18, 0.67), (0.21, 0.67))
    arrow_between(ax2, (0.39, 0.67), (0.42, 0.67))
    arrow_between(ax2, (0.58, 0.67), (0.61, 0.67))
    arrow_between(ax2, (0.69, 0.58), (0.24, 0.37), color="#95A9C6")
    arrow_between(ax2, (0.32, 0.28), (0.4, 0.28))
    arrow_between(ax2, (0.57, 0.28), (0.67, 0.28), color="#97B648")
    ax2.text(
        0.02,
        0.06,
        "Realized population effects are smaller than full-switch household effects because targeting, reach,\n"
        "delivery, uptake, and residual unsafe use all attenuate the biological potential benefit.",
        fontsize=9.5,
        color="#384860",
    )

    fig.tight_layout()
    fig.savefig(output_dir / "figure_1_kitchen_cognition_pathway.png", dpi=220, bbox_inches="tight")
    fig.savefig(output_dir / "figure_1_kitchen_cognition_pathway.pdf", bbox_inches="tight")
    plt.close(fig)


def deterministic_bcr(package, override=None, metric="bcr_cog"):
    p = mode_params()
    if override:
        for name, value in override.items():
            p[name] = np.array([float(value)])
        if "share_prenatal_last4m" in override or "share_year1" in override:
            p["share_years2to5"] = 1.0 - p["share_prenatal_last4m"] - p["share_year1"]
        validate_draws(p)
    df = simulate_kitchen(p) if package == "kitchen" else simulate_kohl(p)
    validate_results(df, f"{package}_deterministic")
    return float(df[metric].iloc[0])


def tornado_data(package, metric="bcr_cog", top_n=TORNADO_TOP_N):
    base_value = deterministic_bcr(package, metric=metric)
    rows = []
    for name, row in PARAMS.iterrows():
        if float(row["min"]) == float(row["max"]):
            continue
        try:
            low = deterministic_bcr(package, {name: float(row["min"])}, metric=metric)
            high = deterministic_bcr(package, {name: float(row["max"])}, metric=metric)
        except ValueError:
            # Some one-at-a-time share overrides create impossible developmental
            # share combinations. We skip those cases rather than forcing an
            # arbitrary projection back into the feasible region.
            continue
        rows.append(
            {
                "variable": name,
                "base_value": base_value,
                "low_value": low,
                "high_value": high,
                "swing": abs(high - low),
                "mode": float(row["mode"]),
                "min": float(row["min"]),
                "max": float(row["max"]),
            }
        )
    tornado = pd.DataFrame(rows).sort_values("swing", ascending=False).head(top_n).reset_index(drop=True)
    return tornado


def plot_tornado(tornado_df, package_name, output_dir, metric_label):
    if plt is None:
        return
    base_value = float(tornado_df["base_value"].iloc[0])
    plot_df = tornado_df.sort_values("swing", ascending=True)
    y = np.arange(len(plot_df))
    lows = np.minimum(plot_df["low_value"], plot_df["high_value"])
    highs = np.maximum(plot_df["low_value"], plot_df["high_value"])

    plt.figure(figsize=(9, 6))
    plt.barh(y, highs - base_value, left=base_value, color="tab:blue", alpha=0.55, label="High")
    plt.barh(y, lows - base_value, left=base_value, color="tab:orange", alpha=0.55, label="Low")
    plt.axvline(base_value, color="black", linestyle="--", linewidth=1.2, label="Mode")
    plt.yticks(y, plot_df["variable"])
    plt.xlabel(metric_label)
    plt.title(f"Tornado Diagram: {package_name} {metric_label}")
    plt.legend()
    plt.tight_layout()
    plt.savefig(output_dir / f"{package_name.lower()}_tornado_{metric_label.lower().replace(' ', '_')}.png", dpi=220)
    plt.close()


def lognormal_params_from_mean_sd(mean, sd):
    sigma2 = np.log(1 + (sd**2) / (mean**2))
    mu = np.log(mean) - sigma2 / 2
    return mu, np.sqrt(sigma2)


def nonlinear_iq_gain(beta, pre_bll, post_bll):
    pre = np.maximum(pre_bll, 0)
    post = np.maximum(post_bll, 0)
    return beta * np.log((pre + 1.0) / (post + 1.0))


def nonlinear_background_robustness(package, other_mean, other_sd, n_draws=50000, seed=RNG_SEED):
    mode_df = simulate_kitchen(mode_params()) if package == "kitchen" else simulate_kohl(mode_params())
    delta_bll = float(mode_df["eff_bll_child_cog_equiv"].iloc[0])
    mu, sigma = lognormal_params_from_mean_sd(other_mean, other_sd)
    rng = np.random.default_rng(seed + (1 if package == "kitchen" else 2))
    other_sources = rng.lognormal(mean=mu, sigma=sigma, size=n_draws)
    pre_bll = other_sources + delta_bll
    post_bll = other_sources
    delta_iq = nonlinear_iq_gain(NONLINEAR_BETA, pre_bll, post_bll)
    return {
        "package": package,
        "other_mean": other_mean,
        "other_sd": other_sd,
        "source_specific_bll_reduction": delta_bll,
        "median_pre_bll": float(np.median(pre_bll)),
        "median_post_bll": float(np.median(post_bll)),
        "median_delta_iq": float(np.median(delta_iq)),
        "mean_delta_iq": float(np.mean(delta_iq)),
    }


def nonlinear_background_bcr_trials(package, other_mean, other_sd, seed=RNG_SEED):
    p = draw_params(seed=seed if package == "kitchen" else seed + 2)
    base_df = simulate_kitchen(p) if package == "kitchen" else simulate_kohl(p)
    mu, sigma = lognormal_params_from_mean_sd(other_mean, other_sd)
    rng = np.random.default_rng(seed + (11 if package == "kitchen" else 12))
    other_sources = rng.lognormal(mean=mu, sigma=sigma, size=len(base_df))
    delta_iq = nonlinear_iq_gain(
        NONLINEAR_BETA,
        other_sources + base_df["eff_bll_child_cog_equiv"].to_numpy(),
        other_sources,
    )
    benefit_earnings = base_df["pv_earnings"].to_numpy() * p["earn_per_iq"] * delta_iq
    return pd.Series(benefit_earnings / base_df["cost"].to_numpy())


def adult_spillover_robustness(package, spillover_share=0.5):
    p = mode_params()
    if package == "kitchen":
        base_df = simulate_kitchen(p)
        base_cvd = float(base_df["benefit_cvd"].iloc[0])
        p_pot = float(base_df["p_success_pot"].iloc[0])
        productivity, _, _, vsly = add_common_economic_fields(p)
        extra_cvd = (
            float(p["bll_pot_mother"][0])
            * p_pot
            * (1 - float(p["residual_unsafe_use_harm"][0]))
            * float(p["cvd_daly_per_ug_lifetime"][0])
            * float(p["years_pot"][0])
            * float(vsly[0])
            * float(np.power(1 + p["discount"][0], -max(0, float(p["adult_reference_years"][0]) - float(p["mom_age"][0]))))
            * spillover_share
        )
    else:
        base_df = simulate_kohl(p)
        base_cvd = float(base_df["benefit_cvd"].iloc[0])
        p_maternal = float(base_df["p_success_kohl_maternal"].iloc[0])
        productivity, _, _, vsly = add_common_economic_fields(p)
        extra_cvd = (
            float(p["bll_maternal_kohl_mother"][0])
            * p_maternal
            * float(p["cvd_daly_per_ug_lifetime"][0])
            * float(p["years_kohl_maternal"][0])
            * float(vsly[0])
            * float(np.power(1 + p["discount"][0], -max(0, float(p["adult_reference_years"][0]) - float(p["mom_age"][0]))))
            * spillover_share
        )
    return {
        "package": package,
        "spillover_share": spillover_share,
        "base_cvd_benefit": base_cvd,
        "spillover_cvd_benefit": base_cvd + extra_cvd,
        "cvd_multiplier": (base_cvd + extra_cvd) / base_cvd if base_cvd > 0 else np.nan,
    }


def adult_spillover_total_bcr_trials(package, spillover_share=0.5, seed=RNG_SEED):
    p = draw_params(seed=seed if package == "kitchen" else seed + 2)
    base_df = simulate_kitchen(p) if package == "kitchen" else simulate_kohl(p)
    _, _, _, vsly = add_common_economic_fields(p)
    if package == "kitchen":
        extra_cvd = (
            base_df["eff_bll_pot_mother"].to_numpy()
            * p["years_pot"]
            * p["cvd_daly_per_ug_lifetime"]
            * vsly
            * np.power(1 + p["discount"], -np.maximum(0, p["adult_reference_years"] - p["mom_age"]))
            * spillover_share
        )
    else:
        extra_cvd = (
            base_df["eff_bll_kohl_mother"].to_numpy()
            * p["years_kohl_maternal"]
            * p["cvd_daly_per_ug_lifetime"]
            * vsly
            * np.power(1 + p["discount"], -np.maximum(0, p["adult_reference_years"] - p["mom_age"]))
            * spillover_share
        )
    total_benefit = (
        base_df["benefit_earnings"].to_numpy()
        + base_df["benefit_maternal_neonatal"].to_numpy()
        + base_df["benefit_cvd"].to_numpy()
        + extra_cvd
    )
    return pd.Series(total_benefit / base_df["cost"].to_numpy())


def example_nonlinear_iq_steps():
    return pd.DataFrame(
        [
            {
                "scenario": "12_to_7",
                "start_bll": 12.0,
                "end_bll": 7.0,
                "delta_iq_nonlinear": nonlinear_iq_gain(NONLINEAR_BETA, 12.0, 7.0),
            },
            {
                "scenario": "7_to_2",
                "start_bll": 7.0,
                "end_bll": 2.0,
                "delta_iq_nonlinear": nonlinear_iq_gain(NONLINEAR_BETA, 7.0, 2.0),
            },
        ]
    )


def kohl_girls_only_robustness(p, kohl_all_infants):
    kohl_mothers_girls = simulate_kohl(p, infant_kohl_multiplier=p["is_daughter"])
    rows = []
    for metric in ["bcr_cog", "bcr_plus_mn", "bcr_total"]:
        all_infants_median = float(kohl_all_infants[metric].median())
        mothers_girls_median = float(kohl_mothers_girls[metric].median())
        rows.append(
            {
                "metric": metric,
                "all_infants_median": all_infants_median,
                "mothers_and_girls_only_median": mothers_girls_median,
                "loss_pct_if_mothers_and_girls_only": (
                    100 * (all_infants_median - mothers_girls_median) / all_infants_median
                    if all_infants_median > 0
                    else np.nan
                ),
                "all_infants_fail_pct": float((kohl_all_infants[metric] < 1).mean() * 100),
                "mothers_and_girls_only_fail_pct": float((kohl_mothers_girls[metric] < 1).mean() * 100),
            }
        )
    return pd.DataFrame(rows), kohl_mothers_girls


def scaled_scenario_trials(package, scenario_params, scale, seed=RNG_SEED):
    original_params = PARAMS.copy(deep=True)
    try:
        globals()["PARAMS"] = scale_param_block(PARAMS, scenario_params, scale)
        if package == "kitchen":
            return simulate_kitchen(draw_params(seed=seed))
        return simulate_kohl(draw_params(seed=seed + 2))
    finally:
        globals()["PARAMS"] = original_params


def export_robustness_distribution_summary(output_dir):
    rows = []

    def add_row(check_name, package, bcr_type, series):
        stats = summarize_results(pd.DataFrame({"metric": series}), "metric")
        rows.append(
            {
                "check": check_name,
                "package": package,
                "bcr_type": bcr_type,
                "median": stats["median"],
                "p5": stats["p5"],
                "p95": stats["p95"],
                "fail_pct": stats["fail_pct"],
            }
        )

    add_row(
        "Adult spillover (+1 adult in half of recipient households)",
        "Kitchen",
        "total",
        adult_spillover_total_bcr_trials("kitchen", spillover_share=0.5),
    )
    add_row(
        "Adult spillover (+1 adult in half of recipient households)",
        "Kohl",
        "total",
        adult_spillover_total_bcr_trials("kohl", spillover_share=0.5),
    )

    add_row(
        "Low-background nonlinear BLL (other-source mean 3, SD 1.5)",
        "Kitchen",
        "cognition",
        nonlinear_background_bcr_trials("kitchen", other_mean=3.0, other_sd=1.5),
    )
    add_row(
        "Low-background nonlinear BLL (other-source mean 3, SD 1.5)",
        "Kohl",
        "cognition",
        nonlinear_background_bcr_trials("kohl", other_mean=3.0, other_sd=1.5),
    )
    add_row(
        "High-background nonlinear BLL (other-source mean 7, SD 3.5)",
        "Kitchen",
        "cognition",
        nonlinear_background_bcr_trials("kitchen", other_mean=7.0, other_sd=3.5),
    )
    add_row(
        "High-background nonlinear BLL (other-source mean 7, SD 3.5)",
        "Kohl",
        "cognition",
        nonlinear_background_bcr_trials("kohl", other_mean=7.0, other_sd=3.5),
    )

    disadvantaged_params = [
        "gdp_ppp_per_capita",
        "growth",
        "p_att_anc",
        "fidelity",
        "p_merchant_stock",
        "p_redeem",
        "adherence",
    ]
    add_row(
        "Disadvantaged implementation setting",
        "Kitchen",
        "total",
        scaled_scenario_trials("kitchen", disadvantaged_params, 0.7)["bcr_total"],
    )
    add_row(
        "Disadvantaged implementation setting",
        "Kohl",
        "total",
        scaled_scenario_trials("kohl", disadvantaged_params, 0.7)["bcr_total"],
    )

    half_effect_params = [
        "bll_pot_child",
        "bll_pot_mother",
        "bll_mug_child",
        "bll_maternal_kohl_mother",
        "bll_infant_kohl_child",
    ]
    add_row(
        "Half cookware and kohl effect sizes",
        "Kitchen",
        "total",
        scaled_scenario_trials("kitchen", half_effect_params, 0.5)["bcr_total"],
    )
    add_row(
        "Half cookware and kohl effect sizes",
        "Kohl",
        "total",
        scaled_scenario_trials("kohl", half_effect_params, 0.5)["bcr_total"],
    )

    p = draw_params(seed=RNG_SEED + 2)
    _, kohl_mothers_girls_trials = kohl_girls_only_robustness(p, simulate_kohl(p))
    add_row(
        "Kohl girls-only infant substitution",
        "Kohl",
        "total",
        kohl_mothers_girls_trials["bcr_total"],
    )

    pd.DataFrame(rows).to_csv(output_dir / "robustness_distribution_summary.csv", index=False)



def export_modes_workbook(output_dir):
    p = mode_params()
    kitchen_mode = simulate_kitchen(p)
    kohl_mode = simulate_kohl(p)
    kohl_girls_mode = simulate_kohl(p, infant_kohl_multiplier=p["is_daughter"])
    mode_df = PARAMS.reset_index()[["variable", "mode", "description", "source_note"]].rename(
        columns={"mode": "mode_value"}
    )
    mode_results = []
    for package_name, df in [
        ("Kitchen", kitchen_mode),
        ("Kohl all infants", kohl_mode),
        ("Kohl mothers and girls only", kohl_girls_mode),
    ]:
        row = {"package": package_name}
        for col in [
            "cost",
            "benefit_earnings",
            "benefit_maternal_neonatal",
            "benefit_cvd",
            "bcr_cog",
            "bcr_plus_mn",
            "bcr_total",
            "eff_bll_child_total",
            "eff_bll_child_cog_equiv",
            "delta_iq",
        ]:
            row[col] = float(df[col].iloc[0])
        mode_results.append(row)
    mode_results_df = pd.DataFrame(mode_results)
    formulas_df = pd.DataFrame(
        [
            {
                "item": "Productivity",
                "formula": "gdp_ppp_per_capita x labor_share_income",
                "note": "Annual market-labour productivity anchor used in PV and VSLY calculations.",
            },
            {
                "item": "PV earnings",
                "formula": "PV = productivity x x^(labor_entry_age) x sum_t [S_t x x^(t-1)]",
                "note": "Where x=(1+growth)/(1+discount) and S_t follows the convex survival path in the Python model.",
            },
            {
                "item": "Kitchen cognitive benefit",
                "formula": "benefit_earnings = PV x earn_per_iq x delta_iq",
                "note": "delta_iq is based on developmentally weighted child BLL reduction.",
            },
            {
                "item": "BCR_Plus_MN",
                "formula": "(benefit_earnings + benefit_maternal_neonatal) / cost",
                "note": "Primary non-CVD extended social return.",
            },
            {
                "item": "BCR_Total",
                "formula": "(benefit_earnings + benefit_maternal_neonatal + benefit_cvd) / cost",
                "note": "Secondary extension adding adult cardiovascular benefits.",
            },
        ]
    )
    refs_df = pd.DataFrame(
        [
            {"topic": "Cookware and lead exposure", "reference": "Weidenhamer et al. 2014; Weidenhamer et al. 2022; Fellows et al. 2025"},
            {"topic": "BLL to IQ", "reference": "Canfield et al. 2003; Lanphear et al. 2005; Jusko et al. 2008; Schnaas et al. 2006"},
            {"topic": "IQ to earnings", "reference": "Ozawa et al. 2022; Grosse and Zhou 2021; Grosse et al. 2002; Hanushek and Woessmann 2008"},
            {"topic": "Prenatal transfer", "reference": "Ettinger et al. 2009; WHO 2024; Vigeh et al. 2023"},
        ]
    )

    workbook_paths = [
        output_dir / "sentinel_firewall_modes.xlsx",
        output_dir / "sentinel_firewall_mode_only_model.xlsx",
    ]
    for workbook_path in workbook_paths:
        with pd.ExcelWriter(workbook_path, engine="openpyxl") as writer:
            mode_df.to_excel(writer, sheet_name="Mode_Parameters", index=False)
            mode_results_df.to_excel(writer, sheet_name="Mode_Results", index=False)
            formulas_df.to_excel(writer, sheet_name="PV_Logic", index=False)
            refs_df.to_excel(writer, sheet_name="Reference_Guide", index=False)

            for sheet_name, df in {
                "Mode_Parameters": mode_df,
                "Mode_Results": mode_results_df,
                "PV_Logic": formulas_df,
                "Reference_Guide": refs_df,
            }.items():
                ws = writer.book[sheet_name]
                for idx, column in enumerate(df.columns, start=1):
                    max_len = max(len(str(column)), *(len(str(v)) for v in df[column].fillna("")))
                    ws.column_dimensions[get_column_letter(idx)].width = min(max_len + 2, 60)
    return workbook_paths[0]


def load_parameter_metadata():
    """Read rationale/citation text from the appendix parameter table."""
    metadata_path = (
        Path(__file__).resolve().parent
        / "submission_package"
        / "supplementary_materials"
        / "parameter_citation_table.md"
    )
    metadata = {}
    if not metadata_path.exists():
        return metadata

    for line in metadata_path.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line.startswith("| `"):
            continue
        parts = [part.strip() for part in line.split("|")[1:-1]]
        if len(parts) < 6:
            continue
        variable = parts[0].strip("`")
        metadata[variable] = {
            "rationale": parts[4],
            "citation": parts[5],
        }
    return metadata


def export_reader_deterministic_workbook(output_dir):
    workbook_path = output_dir / "sentinel_firewall_reader_deterministic_model.xlsx"
    metadata = load_parameter_metadata()
    wb = Workbook()

    title_fill = PatternFill("solid", fgColor="1F4E78")
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    default_fill = PatternFill("solid", fgColor="EDEDED")
    local_fill = PatternFill("solid", fgColor="FFF2CC")
    derived_fill = PatternFill("solid", fgColor="E2F0D9")
    caution_fill = PatternFill("solid", fgColor="FCE4D6")
    policy_fill = PatternFill("solid", fgColor="EAF3FF")
    white_font = Font(color="FFFFFF", bold=True)

    ws = wb.active
    ws.title = "Instructions"
    ws["A1"] = "Sentinel Firewall local adaptation workbook"
    ws["A1"].font = Font(bold=True, size=16)
    ws["A1"].fill = title_fill
    ws["A1"].font = white_font

    instruction_rows = [
        (
            "Purpose",
            "This workbook accompanies the paper 'The Sentinel Firewall: A Health-System Strategy to Decouple Informal Industrialization from Lead Exposure in LMICs.' The paper asks whether maternal and child health platforms can temporarily reduce lead exposure by replacing contaminated household products with safe substitutes while regulation catches up. This spreadsheet is a simplified local-decision version of that model.",
        ),
        (
            "What this tool does",
            "It compares two packages: a Kitchen Package (safe pot, calcium, child utensils) and a Kohl Package (safe maternal and infant kohl). It converts programme delivery into lower maternal and child blood lead levels, child IQ gains, lifetime earnings gains, maternal and neonatal health gains, and secondary adult cardiovascular gains.",
        ),
        (
            "Who should use it",
            "This workbook is designed for Ministries of Health, NGOs, environmental health programmes, and local researchers who want to ask a practical question: if local prices, prevalence, attendance, and implementation quality look like our setting, would targeted safe-product substitution likely be worth funding?",
        ),
        (
            "How to use it",
            "Go to the Inputs sheet. The grey Default column contains the author's modal assumptions and is locked. Enter local values only in the yellow 'Your local value' column. The green 'Active value' column will automatically use your entry when present and otherwise fall back to the default. The Topline sheet then compares the workbook-default mode-only results with your local mode-only scenario.",
        ),
        (
            "What to change first",
            "Start with parameters local decision-makers are most likely to know or influence: GDP per capita, labour share, mortality, product prices, district births and overhead, prevalence in the targeted area, attendance, voucher redemption, stock, and adherence. Scientific parameters such as BLL-to-IQ and IQ-to-earnings are grouped separately and marked 'change cautiously'.",
        ),
        (
            "How productivity is defined",
            "Productivity is measured as GDP per capita at purchasing power parity multiplied by labour's share of national income. This values market labour output, including self-employment and informal production that would be missed by observed wages alone. It excludes household production and the intrinsic value of cognition beyond earnings, so the earnings channel remains conservative on that dimension.",
        ),
        (
            "Important limitation",
            "This is a deterministic mode-only tool. It does not run the 10,000-draw Monte Carlo uncertainty analysis, does not model parameter correlations, and does not export trial-level data. Use the public Python model for the full probabilistic sensitivity analysis and publication-quality figures.",
        ),
        (
            "Public files",
            "Public copies are posted in the Sentinel-Firewall GitHub repository at https://github.com/levine63/Sentinel-Firewall. The main public files are submission_package/public_code_and_data/sentinel_firewall_model.py and submission_package/public_code_and_data/Simulating_sentinel_safe_products.xlsx.",
        ),
    ]
    row_num = 3
    for heading, body in instruction_rows:
        ws.cell(row_num, 1, heading).font = Font(bold=True)
        ws.cell(row_num, 2, body)
        row_num += 2
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 122

    display_names = {
        "gdp_ppp_per_capita": ("GDP per capita, PPP", "US$ per person per year"),
        "labor_share_income": ("Labour share of national income", "Share of GDP"),
        "growth": ("Real annual productivity growth", "Annual rate"),
        "discount": ("Social discount rate", "Annual rate"),
        "p18": ("Survival from birth to age 18", "Probability"),
        "p65": ("Survival from birth to age 65", "Probability"),
        "k": ("Mortality timing shape parameter", "Unitless"),
        "iq_per_bll": ("IQ gain per 1 ug/dL lower developmental blood lead", "IQ points per ug/dL"),
        "earn_per_iq": ("Earnings gain per IQ point", "Share of earnings"),
        "p_att_anc": ("Attendance at antenatal care contact", "Probability"),
        "p_att_imm": ("Attendance at 6-month immunization contact", "Probability"),
        "p_att_birth": ("Institutional delivery contact", "Probability"),
        "p_att_infant_kohl_contact": ("Infant kohl contact through facility delivery or early immunization", "Probability"),
        "fetal_transfer_coeff": ("Maternal-to-fetal lead reduction transfer", "Share"),
        "lactational_transfer_coeff": ("Maternal-to-infant transfer during the milk-dominated first year", "Share"),
        "share_prenatal_last4m": ("Share of cognitive harm in last four months in utero", "Share"),
        "share_year1": ("Share of cognitive harm in year 1", "Share"),
        "prev_pot": ("Unsafe pot prevalence in targeted households", "Share"),
        "prev_kohl_maternal": ("Unsafe maternal kohl prevalence in targeted households", "Share"),
        "prev_kohl_infant": ("Unsafe infant kohl prevalence in targeted households", "Share"),
        "prev_utensils": ("Unsafe child feeding-utensil prevalence in targeted households", "Share"),
        "fidelity": ("Health-system delivery fidelity", "Probability"),
        "adherence": ("Complete substitution or adherence", "Probability"),
        "p_voucher_issued": ("Pot voucher issued at ANC", "Probability"),
        "p_merchant_stock": ("Approved merchant has safe pot in stock", "Probability"),
        "p_redeem": ("Pot voucher redeemed", "Probability"),
        "p_use_targeted_after_redemption": ("Redeemed pot used for pregnant woman or young child", "Probability"),
        "residual_unsafe_use_harm": ("Residual harm from continued unsafe pot use", "Share of pot benefit lost"),
        "labor_entry_age": ("Age when earnings begin", "Years"),
        "mom_age": ("Mother age at intervention", "Years"),
        "gp_age": ("Grandparent age at intervention", "Years"),
        "gp_coresidence_prob": ("Grandparent co-residence", "Probability"),
        "bll_pot_child": ("Child BLL reduction from full switch to safe pot", "ug/dL during active use"),
        "bll_pot_mother": ("Mother BLL reduction from full switch to safe pot", "ug/dL during active use"),
        "bll_mug_child": ("Child BLL reduction from full switch to safe feeding utensils", "ug/dL during active use"),
        "bll_calc_mother": ("Mother BLL reduction from calcium", "ug/dL during active use"),
        "bll_maternal_kohl_mother": ("Mother BLL reduction from safe maternal kohl", "ug/dL during active use"),
        "bll_infant_kohl_child": ("Child BLL reduction from safe infant kohl", "ug/dL during active use"),
        "cost_pot_direct": ("Lead-safe pot product cost", "US$"),
        "cost_utensils_direct": ("Safe child feeding-utensil product cost", "US$"),
        "cost_calcium_direct": ("Calcium course product cost", "US$"),
        "cost_kohl_maternal_direct": ("Safe maternal kohl product cost", "US$"),
        "cost_kohl_infant_direct": ("Safe infant kohl product cost", "US$"),
        "cost_explain": ("Counseling and explanation cost per contact", "US$"),
        "cost_bulk_distribution_margin": ("Bulk distribution plus retailer margin for pot", "US$"),
        "cost_dist_light": ("Distribution cost for small item", "US$"),
        "logistics_markup": ("Programme overhead markup", "Share of direct product plus distribution cost"),
        "district_fixed_cost": ("National and district fixed programme cost per screening unit", "US$"),
        "district_births": ("Births per screening unit", "Births"),
        "preeclampsia_base_prev": ("Baseline preeclampsia prevalence", "Probability"),
        "preeclampsia_or": ("Preeclampsia odds ratio per 1 ug/dL lead", "Odds ratio"),
        "preterm_base_prev": ("Baseline preterm birth prevalence", "Probability"),
        "preterm_or": ("Preterm birth odds ratio per 1 ug/dL lead", "Odds ratio"),
        "preeclampsia_timing_mult": ("Pregnancy timing adjustment for preeclampsia pathway", "Share"),
        "preterm_timing_mult": ("Pregnancy timing adjustment for preterm pathway", "Share"),
        "neonatal_daly_mult": ("DALYs per preterm birth averted", "DALYs"),
        "cvd_daly_per_ug_lifetime": ("Adult cardiovascular DALYs per 1 ug/dL lifetime lead reduction", "DALYs"),
        "vsly_mult": ("Value per DALY relative to annual productivity", "Multiplier"),
        "preeclampsia_daly_wt": ("DALYs per preeclampsia case", "DALYs"),
        "years_pot": ("Years safe pot protects", "Years"),
        "years_kohl_maternal": ("Years safe maternal kohl protects", "Years"),
        "years_kohl_infant": ("Years safe infant kohl protects", "Years"),
        "years_utensils": ("Years safe utensils protect", "Years"),
        "years_calcium": ("Years calcium affects maternal lead", "Years"),
        "adult_reference_years": ("Reference age for adult cardiovascular benefits", "Years"),
        "gp_pot_effect_multiplier": ("Grandparent pot exposure multiplier", "Share of child pot effect"),
    }
    group_map = {
        "gdp_ppp_per_capita": ("Local macro and demography", "Usually known locally or from public statistics"),
        "labor_share_income": ("Local macro and demography", "Usually known from public macro data or left at default"),
        "growth": ("Local macro and demography", "Change if you have a strong local reason"),
        "discount": ("Local macro and demography", "Policy/evaluation choice"),
        "p18": ("Local macro and demography", "Usually known locally or from public statistics"),
        "p65": ("Local macro and demography", "Usually known locally or from public statistics"),
        "k": ("Local macro and demography", "Usually leave at default"),
        "labor_entry_age": ("Local macro and demography", "Usually known locally"),
        "mom_age": ("Local macro and demography", "Usually known locally"),
        "gp_age": ("Local macro and demography", "Usually leave at default"),
        "gp_coresidence_prob": ("Local macro and demography", "Usually known locally"),
        "prev_pot": ("Targeting and prevalence", "Good local input"),
        "prev_kohl_maternal": ("Targeting and prevalence", "Good local input"),
        "prev_kohl_infant": ("Targeting and prevalence", "Good local input"),
        "prev_utensils": ("Targeting and prevalence", "Good local input"),
        "p_att_anc": ("Implementation and programme reach", "Good local input"),
        "p_att_imm": ("Implementation and programme reach", "Good local input"),
        "p_att_birth": ("Implementation and programme reach", "Good local input"),
        "p_att_infant_kohl_contact": ("Implementation and programme reach", "Good local input"),
        "fidelity": ("Implementation and programme reach", "Programme lever"),
        "adherence": ("Implementation and programme reach", "Programme lever"),
        "p_voucher_issued": ("Implementation and programme reach", "Programme lever"),
        "p_merchant_stock": ("Implementation and programme reach", "Programme lever"),
        "p_redeem": ("Implementation and programme reach", "Programme lever"),
        "p_use_targeted_after_redemption": ("Implementation and programme reach", "Programme lever"),
        "residual_unsafe_use_harm": ("Implementation and programme reach", "Programme lever"),
        "cost_pot_direct": ("Prices and overhead", "Good local input"),
        "cost_utensils_direct": ("Prices and overhead", "Good local input"),
        "cost_calcium_direct": ("Prices and overhead", "Good local input"),
        "cost_kohl_maternal_direct": ("Prices and overhead", "Good local input"),
        "cost_kohl_infant_direct": ("Prices and overhead", "Good local input"),
        "cost_explain": ("Prices and overhead", "Programme lever"),
        "cost_bulk_distribution_margin": ("Prices and overhead", "Programme lever"),
        "cost_dist_light": ("Prices and overhead", "Programme lever"),
        "logistics_markup": ("Prices and overhead", "Programme lever"),
        "district_fixed_cost": ("Prices and overhead", "Good local input"),
        "district_births": ("Prices and overhead", "Good local input"),
        "fetal_transfer_coeff": ("Scientific and valuation parameters", "Change cautiously"),
        "lactational_transfer_coeff": ("Scientific and valuation parameters", "Change cautiously"),
        "share_prenatal_last4m": ("Scientific and valuation parameters", "Change cautiously"),
        "share_year1": ("Scientific and valuation parameters", "Change cautiously"),
        "iq_per_bll": ("Scientific and valuation parameters", "Change cautiously"),
        "earn_per_iq": ("Scientific and valuation parameters", "Change cautiously"),
        "preeclampsia_base_prev": ("Secondary health parameters", "Good local input if available"),
        "preeclampsia_or": ("Secondary health parameters", "Change cautiously"),
        "preterm_base_prev": ("Secondary health parameters", "Good local input if available"),
        "preterm_or": ("Secondary health parameters", "Change cautiously"),
        "preeclampsia_timing_mult": ("Secondary health parameters", "Change cautiously"),
        "preterm_timing_mult": ("Secondary health parameters", "Change cautiously"),
        "neonatal_daly_mult": ("Secondary health parameters", "Change cautiously"),
        "cvd_daly_per_ug_lifetime": ("Secondary health parameters", "Change cautiously"),
        "vsly_mult": ("Secondary health parameters", "Policy choice"),
        "preeclampsia_daly_wt": ("Secondary health parameters", "Change cautiously"),
        "bll_pot_child": ("Product-specific BLL effect sizes", "Change cautiously"),
        "bll_pot_mother": ("Product-specific BLL effect sizes", "Change cautiously"),
        "bll_mug_child": ("Product-specific BLL effect sizes", "Change cautiously"),
        "bll_calc_mother": ("Product-specific BLL effect sizes", "Change cautiously"),
        "bll_maternal_kohl_mother": ("Product-specific BLL effect sizes", "Change cautiously"),
        "bll_infant_kohl_child": ("Product-specific BLL effect sizes", "Change cautiously"),
        "years_pot": ("Timing and duration", "Good local input if programme design differs"),
        "years_kohl_maternal": ("Timing and duration", "Good local input if programme design differs"),
        "years_kohl_infant": ("Timing and duration", "Good local input if programme design differs"),
        "years_utensils": ("Timing and duration", "Good local input if programme design differs"),
        "years_calcium": ("Timing and duration", "Good local input if programme design differs"),
        "adult_reference_years": ("Timing and duration", "Usually leave at default"),
        "gp_pot_effect_multiplier": ("Timing and duration", "Usually leave at default"),
    }
    group_order = [
        "Local macro and demography",
        "Prices and overhead",
        "Targeting and prevalence",
        "Implementation and programme reach",
        "Product-specific BLL effect sizes",
        "Scientific and valuation parameters",
        "Secondary health parameters",
        "Timing and duration",
    ]

    inputs = wb.create_sheet("Inputs")
    headers = [
        "Parameter group",
        "Parameter",
        "Units",
        "Default modal value (locked)",
        "Your local value (edit here)",
        "Active value",
        "Low",
        "High",
        "Why it matters / plain-language description",
        "Citation basis",
        "Change guidance",
        "Internal name",
    ]
    inputs.append(headers)
    for cell in inputs[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
    row_lookup = {}
    grouped_variables = {group: [] for group in group_order}
    for name in PARAMS.index:
        group_name, guidance = group_map.get(name, ("Other", "Change cautiously"))
        grouped_variables.setdefault(group_name, []).append((name, guidance))

    current_row = 2
    for group_name in group_order:
        variables = grouped_variables.get(group_name, [])
        if not variables:
            continue
        inputs.cell(current_row, 1, group_name)
        inputs.cell(current_row, 1).font = Font(bold=True)
        inputs.cell(current_row, 1).fill = policy_fill if "Scientific" not in group_name else caution_fill
        current_row += 1
        for name, guidance in variables:
            row = PARAMS.loc[name]
            label, units = display_names.get(name, (name.replace("_", " "), ""))
            citation = metadata.get(name, {}).get("citation", row["source_note"])
            rationale = metadata.get(name, {}).get("rationale", row["description"])
            row_lookup[name] = current_row
            inputs.cell(current_row, 1, group_name)
            inputs.cell(current_row, 2, label)
            inputs.cell(current_row, 3, units)
            inputs.cell(current_row, 4, float(row["mode"]))
            inputs.cell(current_row, 5, "")
            inputs.cell(current_row, 6, f'=IF(LEN(E{current_row})=0,D{current_row},E{current_row})')
            inputs.cell(current_row, 7, float(row["min"]))
            inputs.cell(current_row, 8, float(row["max"]))
            inputs.cell(current_row, 9, rationale)
            inputs.cell(current_row, 10, citation)
            inputs.cell(current_row, 11, guidance)
            inputs.cell(current_row, 12, name)
            inputs.cell(current_row, 4).fill = default_fill
            inputs.cell(current_row, 5).fill = local_fill
            inputs.cell(current_row, 6).fill = derived_fill
            if guidance == "Change cautiously":
                inputs.cell(current_row, 11).fill = caution_fill
            elif guidance in {"Programme lever", "Good local input", "Usually known locally or from public statistics"}:
                inputs.cell(current_row, 11).fill = policy_fill
            current_row += 1

    for col, width in {
        "A": 30,
        "B": 40,
        "C": 20,
        "D": 16,
        "E": 18,
        "F": 18,
        "G": 11,
        "H": 11,
        "I": 74,
        "J": 78,
        "K": 22,
        "L": 22,
    }.items():
        inputs.column_dimensions[col].width = width
    inputs.column_dimensions["L"].hidden = True
    inputs.freeze_panes = "A2"
    inputs.protection.sheet = True
    for row in range(2, current_row):
        inputs.cell(row, 5).protection = Protection(locked=False)

    def ref_default(name):
        return f"Inputs!$D${row_lookup[name]}"

    def ref_local(name):
        return f"Inputs!$F${row_lookup[name]}"

    guide = wb.create_sheet("Formula_Guide")
    guide.append(["Concept", "Plain-language meaning", "Formula used in this workbook", "Notes"])
    for cell in guide[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
    guide_rows = [
        (
            "Present value of lifetime earnings",
            "Expected discounted lifetime earnings from birth.",
            "Productivity x growth/discount adjustment from birth to labour entry x sum of working-year survival and growth/discount terms.",
            "Productivity is GDP PPP per capita x labour share of national income.",
        ),
        (
            "Productivity",
            "Annual market-labour output used to value lifetime earnings and VSLY.",
            "GDP per capita, PPP x labour share of national income.",
            "This follows social cost-benefit practice and includes self-employment and informal market production better than observed wages alone.",
        ),
        (
            "Pot implementation chain",
            "The safe pot only generates benefit when the household is correctly targeted, reaches ANC, receives a voucher, finds a stocked merchant, redeems the voucher, and uses the pot for the pregnant woman or young child.",
            "Unsafe-pot prevalence x ANC attendance x voucher issued x merchant stocked x voucher redeemed x targeted use x (1 - residual unsafe-use harm).",
            "Costs for the pot are incurred only after attendance, voucher issuance, merchant stock, and redemption; fixed overhead remains regardless.",
        ),
        (
            "Clinic-distributed items",
            "Calcium, child utensils, maternal kohl, and infant kohl use shorter delivery chains.",
            "Relevant prevalence where applicable x relevant health-system contact x fidelity x adherence.",
            "Calcium currently helps all reached pregnant women in the targeted area rather than only women with a measured lead source.",
        ),
        (
            "Developmental timing",
            "An intervention that lasts only during part of pregnancy or early childhood does not receive the full cognitive benefit of lowering BLL throughout the developmental window.",
            "Prenatal BLL reduction x prenatal share + postnatal BLL reduction x years-protected share.",
            "The default shares are 30% for the last four months in utero, 30% for year 1, and 40% for years 2-5 combined.",
        ),
        (
            "IQ gain",
            "Lower developmentally weighted child blood lead increases expected IQ.",
            "Developmentally weighted child BLL reduction x IQ gain per 1 ug/dL lower BLL.",
            "The full Python model also reports a nonlinear log-BLL robustness check.",
        ),
        (
            "Lifetime earnings gain",
            "IQ gains raise expected lifetime earnings.",
            "Present value of lifetime earnings x earnings gain per IQ point x IQ gain.",
            "This is the main cognition-channel social benefit.",
        ),
        (
            "Benefit-cost ratio",
            "Social benefits divided by programme costs.",
            "Benefits / costs.",
            "The Topline sheet compares the workbook-default mode-only run with your local mode-only run for cognition-only, maternal/neonatal, and total BCRs.",
        ),
    ]
    for row in guide_rows:
        guide.append(row)
    for col, width in {"A": 28, "B": 56, "C": 76, "D": 64}.items():
        guide.column_dimensions[col].width = width

    pv = wb.create_sheet("PV_Helper")
    pv_headers = [
        "Working year",
        "Survival S_t (default)",
        "Survival S_t (local)",
        "Growth-discount factor (default)",
        "Growth-discount factor (local)",
        "Included in work life (default)",
        "Included in work life (local)",
        "PV term (default)",
        "PV term (local)",
    ]
    pv.append(pv_headers)
    for cell in pv[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
    x_formula_default = f"((1+{ref_default('growth')})/(1+{ref_default('discount')}))"
    x_formula_local = f"((1+{ref_local('growth')})/(1+{ref_local('discount')}))"
    t_work_default = f"(65-{ref_default('labor_entry_age')}+1)"
    t_work_local = f"(65-{ref_local('labor_entry_age')}+1)"
    for i in range(1, 53):
        r = i + 1
        pv.cell(r, 1, i)
        pv.cell(r, 2, f"={ref_default('p18')}-({ref_default('p18')}-{ref_default('p65')})*POWER((A{r}-1)/47,{ref_default('k')})")
        pv.cell(r, 3, f"={ref_local('p18')}-({ref_local('p18')}-{ref_local('p65')})*POWER((A{r}-1)/47,{ref_local('k')})")
        pv.cell(r, 4, f"=POWER({x_formula_default},A{r}-1)")
        pv.cell(r, 5, f"=POWER({x_formula_local},A{r}-1)")
        pv.cell(r, 6, f"=--(A{r}<={t_work_default})")
        pv.cell(r, 7, f"=--(A{r}<={t_work_local})")
        pv.cell(r, 8, f"=B{r}*D{r}*F{r}")
        pv.cell(r, 9, f"=C{r}*E{r}*G{r}")
    for col in range(1, 10):
        pv.column_dimensions[get_column_letter(col)].width = 22

    results = wb.create_sheet("Results")
    results["A1"] = "Detailed deterministic calculations"
    results["A1"].font = Font(bold=True, size=14)
    results["A2"] = "Default paper values"
    results["B2"].fill = default_fill
    results["B2"] = "Default"
    results["C2"].fill = derived_fill
    results["C2"] = "Your local run"
    results["D2"] = "Notes"
    for cell in results["A2:D2"][0]:
        cell.font = Font(bold=True)
        if cell.column == 1 or cell.column == 4:
            cell.fill = header_fill
    for col, width in {"A": 36, "B": 16, "C": 16, "D": 60}.items():
        results.column_dimensions[col].width = width
    rows = {}

    def res_ref(label, local=False):
        col = "C" if local else "B"
        return f"Results!${col}${rows[label]}"

    def add_calc(label, formula_default=None, formula_local=None, note=None):
        row = results.max_row + 1
        results.cell(row, 1, label)
        if formula_default is not None:
            results.cell(row, 2, formula_default)
        if formula_local is not None:
            results.cell(row, 3, formula_local)
        if note is not None:
            results.cell(row, 4, note)
        rows[label] = row
        return row

    add_calc(
        "productivity",
        f"={ref_default('gdp_ppp_per_capita')}*{ref_default('labor_share_income')}",
        f"={ref_local('gdp_ppp_per_capita')}*{ref_local('labor_share_income')}",
        "GDP PPP per capita x labour share of national income",
    )
    add_calc(
        "x_growth_discount",
        f"={x_formula_default}",
        f"={x_formula_local}",
        "(1+growth)/(1+discount)",
    )
    add_calc(
        "pv_earnings",
        f"=B{rows['productivity']}*POWER(B{rows['x_growth_discount']},{ref_default('labor_entry_age')})*SUM(PV_Helper!H2:H53)",
        f"=C{rows['productivity']}*POWER(C{rows['x_growth_discount']},{ref_local('labor_entry_age')})*SUM(PV_Helper!I2:I53)",
    )
    add_calc(
        "fixed_cost_per_birth",
        f"={ref_default('district_fixed_cost')}/{ref_default('district_births')}",
        f"={ref_local('district_fixed_cost')}/{ref_local('district_births')}",
    )
    add_calc(
        "vsly",
        f"={ref_default('vsly_mult')}*B{rows['productivity']}",
        f"={ref_local('vsly_mult')}*C{rows['productivity']}",
    )
    add_calc(
        "share_years2to5",
        f"=1-{ref_default('share_prenatal_last4m')}-{ref_default('share_year1')}",
        f"=1-{ref_local('share_prenatal_last4m')}-{ref_local('share_year1')}",
    )
    add_calc(
        "post_weight_pot",
        f"={ref_default('share_year1')}*MIN({ref_default('years_pot')},0.5)+B{rows['share_years2to5']}*(MIN(MAX({ref_default('years_pot')}-0.5,0),4)/4)",
        f"={ref_local('share_year1')}*MIN({ref_local('years_pot')},0.5)+C{rows['share_years2to5']}*(MIN(MAX({ref_local('years_pot')}-0.5,0),4)/4)",
    )
    add_calc(
        "post_weight_utensils",
        f"={ref_default('share_year1')}*MIN({ref_default('years_utensils')},0.5)+B{rows['share_years2to5']}*(MIN(MAX({ref_default('years_utensils')}-0.5,0),4)/4)",
        f"={ref_local('share_year1')}*MIN({ref_local('years_utensils')},0.5)+C{rows['share_years2to5']}*(MIN(MAX({ref_local('years_utensils')}-0.5,0),4)/4)",
    )
    add_calc(
        "post_weight_kohl_infant",
        f"={ref_default('share_year1')}*MIN({ref_default('years_kohl_infant')},1)+B{rows['share_years2to5']}*(MIN(MAX({ref_default('years_kohl_infant')}-1,0),4)/4)",
        f"={ref_local('share_year1')}*MIN({ref_local('years_kohl_infant')},1)+C{rows['share_years2to5']}*(MIN(MAX({ref_local('years_kohl_infant')}-1,0),4)/4)",
    )
    add_calc("")
    add_calc("Kitchen package", note="Default and local deterministic runs")
    results.cell(rows["Kitchen package"], 1).font = Font(bold=True)
    add_calc(
        "k_p_success_pot",
        f"={ref_default('prev_pot')}*{ref_default('p_att_anc')}*{ref_default('p_voucher_issued')}*{ref_default('p_merchant_stock')}*{ref_default('p_redeem')}*{ref_default('p_use_targeted_after_redemption')}",
        f"={ref_local('prev_pot')}*{ref_local('p_att_anc')}*{ref_local('p_voucher_issued')}*{ref_local('p_merchant_stock')}*{ref_local('p_redeem')}*{ref_local('p_use_targeted_after_redemption')}",
    )
    add_calc(
        "k_p_success_pot_after_residual",
        f"=B{rows['k_p_success_pot']}*(1-{ref_default('residual_unsafe_use_harm')})",
        f"=C{rows['k_p_success_pot']}*(1-{ref_local('residual_unsafe_use_harm')})",
    )
    add_calc(
        "k_p_success_calcium",
        f"={ref_default('p_att_anc')}*{ref_default('fidelity')}*{ref_default('adherence')}",
        f"={ref_local('p_att_anc')}*{ref_local('fidelity')}*{ref_local('adherence')}",
    )
    add_calc(
        "k_p_success_utensils",
        f"={ref_default('prev_utensils')}*{ref_default('p_att_imm')}*{ref_default('fidelity')}*{ref_default('adherence')}",
        f"={ref_local('prev_utensils')}*{ref_local('p_att_imm')}*{ref_local('fidelity')}*{ref_local('adherence')}",
    )
    add_calc(
        "k_eff_bll_pot_child",
        f"={ref_default('bll_pot_child')}*B{rows['k_p_success_pot_after_residual']}",
        f"={ref_local('bll_pot_child')}*C{rows['k_p_success_pot_after_residual']}",
    )
    add_calc(
        "k_eff_bll_pot_mother",
        f"={ref_default('bll_pot_mother')}*B{rows['k_p_success_pot_after_residual']}",
        f"={ref_local('bll_pot_mother')}*C{rows['k_p_success_pot_after_residual']}",
    )
    add_calc(
        "k_eff_bll_calcium_mother",
        f"={ref_default('bll_calc_mother')}*B{rows['k_p_success_calcium']}",
        f"={ref_local('bll_calc_mother')}*C{rows['k_p_success_calcium']}",
    )
    add_calc(
        "k_eff_bll_utensils_child",
        f"={ref_default('bll_mug_child')}*B{rows['k_p_success_utensils']}",
        f"={ref_local('bll_mug_child')}*C{rows['k_p_success_utensils']}",
    )
    add_calc(
        "k_eff_bll_prenatal_child",
        f"={ref_default('fetal_transfer_coeff')}*(B{rows['k_eff_bll_pot_mother']}+B{rows['k_eff_bll_calcium_mother']})",
        f"={ref_local('fetal_transfer_coeff')}*(C{rows['k_eff_bll_pot_mother']}+C{rows['k_eff_bll_calcium_mother']})",
    )
    add_calc(
        "k_eff_bll_lactational_child",
        f"={ref_default('lactational_transfer_coeff')}*B{rows['k_eff_bll_pot_mother']}",
        f"={ref_local('lactational_transfer_coeff')}*C{rows['k_eff_bll_pot_mother']}",
    )
    add_calc(
        "k_eff_bll_child_total",
        f"=B{rows['k_eff_bll_pot_child']}+B{rows['k_eff_bll_utensils_child']}+B{rows['k_eff_bll_prenatal_child']}+B{rows['k_eff_bll_lactational_child']}",
        f"=C{rows['k_eff_bll_pot_child']}+C{rows['k_eff_bll_utensils_child']}+C{rows['k_eff_bll_prenatal_child']}+C{rows['k_eff_bll_lactational_child']}",
    )
    add_calc(
        "k_eff_bll_child_cog_equiv",
        f"=B{rows['k_eff_bll_prenatal_child']}*{ref_default('share_prenatal_last4m')}+B{rows['k_eff_bll_lactational_child']}*{ref_default('share_year1')}*MIN({ref_default('years_pot')},1)+B{rows['k_eff_bll_pot_child']}*B{rows['post_weight_pot']}+B{rows['k_eff_bll_utensils_child']}*B{rows['post_weight_utensils']}",
        f"=C{rows['k_eff_bll_prenatal_child']}*{ref_local('share_prenatal_last4m')}+C{rows['k_eff_bll_lactational_child']}*{ref_local('share_year1')}*MIN({ref_local('years_pot')},1)+C{rows['k_eff_bll_pot_child']}*C{rows['post_weight_pot']}+C{rows['k_eff_bll_utensils_child']}*C{rows['post_weight_utensils']}",
    )
    add_calc(
        "k_delta_iq",
        f"=B{rows['k_eff_bll_child_cog_equiv']}*{ref_default('iq_per_bll')}",
        f"=C{rows['k_eff_bll_child_cog_equiv']}*{ref_local('iq_per_bll')}",
    )
    add_calc(
        "k_benefit_earnings",
        f"=B{rows['pv_earnings']}*{ref_default('earn_per_iq')}*B{rows['k_delta_iq']}",
        f"=C{rows['pv_earnings']}*{ref_local('earn_per_iq')}*C{rows['k_delta_iq']}",
    )
    add_calc(
        "k_preeclampsia_delta",
        f"={ref_default('preeclampsia_base_prev')}-(({ref_default('preeclampsia_base_prev')}/(1-{ref_default('preeclampsia_base_prev')}))/POWER({ref_default('preeclampsia_or')},(B{rows['k_eff_bll_pot_mother']}+B{rows['k_eff_bll_calcium_mother']})*{ref_default('preeclampsia_timing_mult')}))/(1+(({ref_default('preeclampsia_base_prev')}/(1-{ref_default('preeclampsia_base_prev')}))/POWER({ref_default('preeclampsia_or')},(B{rows['k_eff_bll_pot_mother']}+B{rows['k_eff_bll_calcium_mother']})*{ref_default('preeclampsia_timing_mult')})))",
        f"={ref_local('preeclampsia_base_prev')}-(({ref_local('preeclampsia_base_prev')}/(1-{ref_local('preeclampsia_base_prev')}))/POWER({ref_local('preeclampsia_or')},(C{rows['k_eff_bll_pot_mother']}+C{rows['k_eff_bll_calcium_mother']})*{ref_local('preeclampsia_timing_mult')}))/(1+(({ref_local('preeclampsia_base_prev')}/(1-{ref_local('preeclampsia_base_prev')}))/POWER({ref_local('preeclampsia_or')},(C{rows['k_eff_bll_pot_mother']}+C{rows['k_eff_bll_calcium_mother']})*{ref_local('preeclampsia_timing_mult')})))",
    )
    add_calc(
        "k_preterm_delta",
        f"={ref_default('preterm_base_prev')}-(({ref_default('preterm_base_prev')}/(1-{ref_default('preterm_base_prev')}))/POWER({ref_default('preterm_or')},(B{rows['k_eff_bll_pot_mother']}+B{rows['k_eff_bll_calcium_mother']})*{ref_default('preterm_timing_mult')}))/(1+(({ref_default('preterm_base_prev')}/(1-{ref_default('preterm_base_prev')}))/POWER({ref_default('preterm_or')},(B{rows['k_eff_bll_pot_mother']}+B{rows['k_eff_bll_calcium_mother']})*{ref_default('preterm_timing_mult')})))",
        f"={ref_local('preterm_base_prev')}-(({ref_local('preterm_base_prev')}/(1-{ref_local('preterm_base_prev')}))/POWER({ref_local('preterm_or')},(C{rows['k_eff_bll_pot_mother']}+C{rows['k_eff_bll_calcium_mother']})*{ref_local('preterm_timing_mult')}))/(1+(({ref_local('preterm_base_prev')}/(1-{ref_local('preterm_base_prev')}))/POWER({ref_local('preterm_or')},(C{rows['k_eff_bll_pot_mother']}+C{rows['k_eff_bll_calcium_mother']})*{ref_local('preterm_timing_mult')})))",
    )
    add_calc(
        "k_benefit_maternal_neonatal",
        f"=B{rows['k_preeclampsia_delta']}*{ref_default('preeclampsia_daly_wt')}*B{rows['vsly']}+B{rows['k_preterm_delta']}*{ref_default('neonatal_daly_mult')}*B{rows['vsly']}",
        f"=C{rows['k_preeclampsia_delta']}*{ref_local('preeclampsia_daly_wt')}*C{rows['vsly']}+C{rows['k_preterm_delta']}*{ref_local('neonatal_daly_mult')}*C{rows['vsly']}",
    )
    add_calc(
        "k_benefit_cvd",
        f"=((B{rows['k_eff_bll_pot_mother']}*{ref_default('years_pot')}+B{rows['k_eff_bll_calcium_mother']}*{ref_default('years_calcium')})*{ref_default('cvd_daly_per_ug_lifetime')}*B{rows['vsly']}*POWER(1+{ref_default('discount')},-MAX(0,{ref_default('adult_reference_years')}-{ref_default('mom_age')})))+(B{rows['k_eff_bll_pot_child']}*{ref_default('cvd_daly_per_ug_lifetime')}*{ref_default('years_pot')}*B{rows['vsly']}*{ref_default('gp_pot_effect_multiplier')}*POWER(1+{ref_default('discount')},-MAX(0,{ref_default('adult_reference_years')}-{ref_default('gp_age')}))*{ref_default('gp_coresidence_prob')})",
        f"=((C{rows['k_eff_bll_pot_mother']}*{ref_local('years_pot')}+C{rows['k_eff_bll_calcium_mother']}*{ref_local('years_calcium')})*{ref_local('cvd_daly_per_ug_lifetime')}*C{rows['vsly']}*POWER(1+{ref_local('discount')},-MAX(0,{ref_local('adult_reference_years')}-{ref_local('mom_age')})))+(C{rows['k_eff_bll_pot_child']}*{ref_local('cvd_daly_per_ug_lifetime')}*{ref_local('years_pot')}*C{rows['vsly']}*{ref_local('gp_pot_effect_multiplier')}*POWER(1+{ref_local('discount')},-MAX(0,{ref_local('adult_reference_years')}-{ref_local('gp_age')}))*{ref_local('gp_coresidence_prob')})",
    )
    add_calc(
        "k_cost",
        f"=B{rows['fixed_cost_per_birth']}+{ref_default('cost_explain')}*{ref_default('p_att_anc')}+({ref_default('cost_pot_direct')}+{ref_default('cost_bulk_distribution_margin')})*(1+{ref_default('logistics_markup')})*{ref_default('p_att_anc')}*{ref_default('p_voucher_issued')}*{ref_default('p_merchant_stock')}*{ref_default('p_redeem')}+{ref_default('cost_explain')}*{ref_default('p_att_anc')}+{ref_default('cost_dist_light')}*(1+{ref_default('logistics_markup')})*{ref_default('p_att_anc')}*{ref_default('fidelity')}+{ref_default('cost_calcium_direct')}*(1+{ref_default('logistics_markup')})*{ref_default('p_att_anc')}*{ref_default('fidelity')}+{ref_default('cost_explain')}*{ref_default('p_att_imm')}+{ref_default('cost_dist_light')}*(1+{ref_default('logistics_markup')})*{ref_default('p_att_imm')}*{ref_default('fidelity')}+{ref_default('cost_utensils_direct')}*(1+{ref_default('logistics_markup')})*{ref_default('p_att_imm')}*{ref_default('fidelity')}",
        f"=C{rows['fixed_cost_per_birth']}+{ref_local('cost_explain')}*{ref_local('p_att_anc')}+({ref_local('cost_pot_direct')}+{ref_local('cost_bulk_distribution_margin')})*(1+{ref_local('logistics_markup')})*{ref_local('p_att_anc')}*{ref_local('p_voucher_issued')}*{ref_local('p_merchant_stock')}*{ref_local('p_redeem')}+{ref_local('cost_explain')}*{ref_local('p_att_anc')}+{ref_local('cost_dist_light')}*(1+{ref_local('logistics_markup')})*{ref_local('p_att_anc')}*{ref_local('fidelity')}+{ref_local('cost_calcium_direct')}*(1+{ref_local('logistics_markup')})*{ref_local('p_att_anc')}*{ref_local('fidelity')}+{ref_local('cost_explain')}*{ref_local('p_att_imm')}+{ref_local('cost_dist_light')}*(1+{ref_local('logistics_markup')})*{ref_local('p_att_imm')}*{ref_local('fidelity')}+{ref_local('cost_utensils_direct')}*(1+{ref_local('logistics_markup')})*{ref_local('p_att_imm')}*{ref_local('fidelity')}",
    )
    add_calc("k_bcr_cog", f"=B{rows['k_benefit_earnings']}/B{rows['k_cost']}", f"=C{rows['k_benefit_earnings']}/C{rows['k_cost']}")
    add_calc("k_bcr_plus_mn", f"=(B{rows['k_benefit_earnings']}+B{rows['k_benefit_maternal_neonatal']})/B{rows['k_cost']}", f"=(C{rows['k_benefit_earnings']}+C{rows['k_benefit_maternal_neonatal']})/C{rows['k_cost']}")
    add_calc("k_bcr_total", f"=(B{rows['k_benefit_earnings']}+B{rows['k_benefit_maternal_neonatal']}+B{rows['k_benefit_cvd']})/B{rows['k_cost']}", f"=(C{rows['k_benefit_earnings']}+C{rows['k_benefit_maternal_neonatal']}+C{rows['k_benefit_cvd']})/C{rows['k_cost']}")
    add_calc("")
    add_calc("Kohl package", note="Default and local deterministic runs")
    results.cell(rows["Kohl package"], 1).font = Font(bold=True)
    add_calc("ko_p_success_maternal", f"={ref_default('prev_kohl_maternal')}*{ref_default('p_att_anc')}*{ref_default('fidelity')}*{ref_default('adherence')}", f"={ref_local('prev_kohl_maternal')}*{ref_local('p_att_anc')}*{ref_local('fidelity')}*{ref_local('adherence')}")
    add_calc("ko_p_success_infant", f"={ref_default('prev_kohl_infant')}*{ref_default('p_att_infant_kohl_contact')}*{ref_default('fidelity')}*{ref_default('adherence')}", f"={ref_local('prev_kohl_infant')}*{ref_local('p_att_infant_kohl_contact')}*{ref_local('fidelity')}*{ref_local('adherence')}")
    add_calc("ko_eff_bll_mother", f"={ref_default('bll_maternal_kohl_mother')}*B{rows['ko_p_success_maternal']}", f"={ref_local('bll_maternal_kohl_mother')}*C{rows['ko_p_success_maternal']}")
    add_calc("ko_eff_bll_infant", f"={ref_default('bll_infant_kohl_child')}*B{rows['ko_p_success_infant']}", f"={ref_local('bll_infant_kohl_child')}*C{rows['ko_p_success_infant']}")
    add_calc("ko_eff_bll_prenatal_child", f"={ref_default('fetal_transfer_coeff')}*B{rows['ko_eff_bll_mother']}", f"={ref_local('fetal_transfer_coeff')}*C{rows['ko_eff_bll_mother']}")
    add_calc("ko_eff_bll_lactational_child", f"={ref_default('lactational_transfer_coeff')}*B{rows['ko_eff_bll_mother']}", f"={ref_local('lactational_transfer_coeff')}*C{rows['ko_eff_bll_mother']}")
    add_calc("ko_eff_bll_child_total", f"=B{rows['ko_eff_bll_prenatal_child']}+B{rows['ko_eff_bll_lactational_child']}+B{rows['ko_eff_bll_infant']}", f"=C{rows['ko_eff_bll_prenatal_child']}+C{rows['ko_eff_bll_lactational_child']}+C{rows['ko_eff_bll_infant']}")
    add_calc("ko_eff_bll_child_cog_equiv", f"=B{rows['ko_eff_bll_prenatal_child']}*{ref_default('share_prenatal_last4m')}+B{rows['ko_eff_bll_lactational_child']}*{ref_default('share_year1')}*MIN({ref_default('years_kohl_maternal')},1)+B{rows['ko_eff_bll_infant']}*B{rows['post_weight_kohl_infant']}", f"=C{rows['ko_eff_bll_prenatal_child']}*{ref_local('share_prenatal_last4m')}+C{rows['ko_eff_bll_lactational_child']}*{ref_local('share_year1')}*MIN({ref_local('years_kohl_maternal')},1)+C{rows['ko_eff_bll_infant']}*C{rows['post_weight_kohl_infant']}")
    add_calc("ko_delta_iq", f"=B{rows['ko_eff_bll_child_cog_equiv']}*{ref_default('iq_per_bll')}", f"=C{rows['ko_eff_bll_child_cog_equiv']}*{ref_local('iq_per_bll')}")
    add_calc("ko_benefit_earnings", f"=B{rows['pv_earnings']}*{ref_default('earn_per_iq')}*B{rows['ko_delta_iq']}", f"=C{rows['pv_earnings']}*{ref_local('earn_per_iq')}*C{rows['ko_delta_iq']}")
    add_calc("ko_preeclampsia_delta", f"={ref_default('preeclampsia_base_prev')}-(({ref_default('preeclampsia_base_prev')}/(1-{ref_default('preeclampsia_base_prev')}))/POWER({ref_default('preeclampsia_or')},B{rows['ko_eff_bll_mother']}*{ref_default('preeclampsia_timing_mult')}))/(1+(({ref_default('preeclampsia_base_prev')}/(1-{ref_default('preeclampsia_base_prev')}))/POWER({ref_default('preeclampsia_or')},B{rows['ko_eff_bll_mother']}*{ref_default('preeclampsia_timing_mult')})))", f"={ref_local('preeclampsia_base_prev')}-(({ref_local('preeclampsia_base_prev')}/(1-{ref_local('preeclampsia_base_prev')}))/POWER({ref_local('preeclampsia_or')},C{rows['ko_eff_bll_mother']}*{ref_local('preeclampsia_timing_mult')}))/(1+(({ref_local('preeclampsia_base_prev')}/(1-{ref_local('preeclampsia_base_prev')}))/POWER({ref_local('preeclampsia_or')},C{rows['ko_eff_bll_mother']}*{ref_local('preeclampsia_timing_mult')})))")
    add_calc("ko_preterm_delta", f"={ref_default('preterm_base_prev')}-(({ref_default('preterm_base_prev')}/(1-{ref_default('preterm_base_prev')}))/POWER({ref_default('preterm_or')},B{rows['ko_eff_bll_mother']}*{ref_default('preterm_timing_mult')}))/(1+(({ref_default('preterm_base_prev')}/(1-{ref_default('preterm_base_prev')}))/POWER({ref_default('preterm_or')},B{rows['ko_eff_bll_mother']}*{ref_default('preterm_timing_mult')})))", f"={ref_local('preterm_base_prev')}-(({ref_local('preterm_base_prev')}/(1-{ref_local('preterm_base_prev')}))/POWER({ref_local('preterm_or')},C{rows['ko_eff_bll_mother']}*{ref_local('preterm_timing_mult')}))/(1+(({ref_local('preterm_base_prev')}/(1-{ref_local('preterm_base_prev')}))/POWER({ref_local('preterm_or')},C{rows['ko_eff_bll_mother']}*{ref_local('preterm_timing_mult')})))")
    add_calc("ko_benefit_maternal_neonatal", f"=B{rows['ko_preeclampsia_delta']}*{ref_default('preeclampsia_daly_wt')}*B{rows['vsly']}+B{rows['ko_preterm_delta']}*{ref_default('neonatal_daly_mult')}*B{rows['vsly']}", f"=C{rows['ko_preeclampsia_delta']}*{ref_local('preeclampsia_daly_wt')}*C{rows['vsly']}+C{rows['ko_preterm_delta']}*{ref_local('neonatal_daly_mult')}*C{rows['vsly']}")
    add_calc("ko_benefit_cvd", f"=B{rows['ko_eff_bll_mother']}*{ref_default('cvd_daly_per_ug_lifetime')}*{ref_default('years_kohl_maternal')}*B{rows['vsly']}*POWER(1+{ref_default('discount')},-MAX(0,{ref_default('adult_reference_years')}-{ref_default('mom_age')}))", f"=C{rows['ko_eff_bll_mother']}*{ref_local('cvd_daly_per_ug_lifetime')}*{ref_local('years_kohl_maternal')}*C{rows['vsly']}*POWER(1+{ref_local('discount')},-MAX(0,{ref_local('adult_reference_years')}-{ref_local('mom_age')}))")
    add_calc("ko_cost", f"=B{rows['fixed_cost_per_birth']}+{ref_default('cost_explain')}*{ref_default('p_att_anc')}+{ref_default('cost_dist_light')}*(1+{ref_default('logistics_markup')})*{ref_default('p_att_anc')}*{ref_default('fidelity')}+{ref_default('cost_kohl_maternal_direct')}*{ref_default('years_kohl_maternal')}*(1+{ref_default('logistics_markup')})*{ref_default('p_att_anc')}*{ref_default('fidelity')}+{ref_default('cost_explain')}*{ref_default('p_att_infant_kohl_contact')}+{ref_default('cost_dist_light')}*(1+{ref_default('logistics_markup')})*{ref_default('p_att_infant_kohl_contact')}*{ref_default('fidelity')}+{ref_default('cost_kohl_infant_direct')}*(1+{ref_default('logistics_markup')})*{ref_default('p_att_infant_kohl_contact')}*{ref_default('fidelity')}", f"=C{rows['fixed_cost_per_birth']}+{ref_local('cost_explain')}*{ref_local('p_att_anc')}+{ref_local('cost_dist_light')}*(1+{ref_local('logistics_markup')})*{ref_local('p_att_anc')}*{ref_local('fidelity')}+{ref_local('cost_kohl_maternal_direct')}*{ref_local('years_kohl_maternal')}*(1+{ref_local('logistics_markup')})*{ref_local('p_att_anc')}*{ref_local('fidelity')}+{ref_local('cost_explain')}*{ref_local('p_att_infant_kohl_contact')}+{ref_local('cost_dist_light')}*(1+{ref_local('logistics_markup')})*{ref_local('p_att_infant_kohl_contact')}*{ref_local('fidelity')}+{ref_local('cost_kohl_infant_direct')}*(1+{ref_local('logistics_markup')})*{ref_local('p_att_infant_kohl_contact')}*{ref_local('fidelity')}")
    add_calc("ko_bcr_cog", f"=B{rows['ko_benefit_earnings']}/B{rows['ko_cost']}", f"=C{rows['ko_benefit_earnings']}/C{rows['ko_cost']}")
    add_calc("ko_bcr_plus_mn", f"=(B{rows['ko_benefit_earnings']}+B{rows['ko_benefit_maternal_neonatal']})/B{rows['ko_cost']}", f"=(C{rows['ko_benefit_earnings']}+C{rows['ko_benefit_maternal_neonatal']})/C{rows['ko_cost']}")
    add_calc("ko_bcr_total", f"=(B{rows['ko_benefit_earnings']}+B{rows['ko_benefit_maternal_neonatal']}+B{rows['ko_benefit_cvd']})/B{rows['ko_cost']}", f"=(C{rows['ko_benefit_earnings']}+C{rows['ko_benefit_maternal_neonatal']}+C{rows['ko_benefit_cvd']})/C{rows['ko_cost']}")

    top = wb.create_sheet("Topline")
    top.append(["Output", "Units", "Paper default: Kitchen", "Your local run: Kitchen", "Paper default: Kohl", "Your local run: Kohl"])
    for cell in top[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
    topline_rows = [
        ("Cost per mother-child pair", "US$", f"={res_ref('k_cost')}", f"={res_ref('k_cost', True)}", f"={res_ref('ko_cost')}", f"={res_ref('ko_cost', True)}"),
        ("Present value of lifetime earnings gain per child", "US$", f"={res_ref('k_benefit_earnings')}", f"={res_ref('k_benefit_earnings', True)}", f"={res_ref('ko_benefit_earnings')}", f"={res_ref('ko_benefit_earnings', True)}"),
        ("Benefit-cost ratio, cognition channel only", "Ratio", f"={res_ref('k_bcr_cog')}", f"={res_ref('k_bcr_cog', True)}", f"={res_ref('ko_bcr_cog')}", f"={res_ref('ko_bcr_cog', True)}"),
        ("Benefit-cost ratio, cognition, maternal health, and neonatal health channels", "Ratio", f"={res_ref('k_bcr_plus_mn')}", f"={res_ref('k_bcr_plus_mn', True)}", f"={res_ref('ko_bcr_plus_mn')}", f"={res_ref('ko_bcr_plus_mn', True)}"),
        ("Benefit-cost ratio, cognition, maternal health, neonatal health, and cardiovascular illness channels", "Ratio", f"={res_ref('k_bcr_total')}", f"={res_ref('k_bcr_total', True)}", f"={res_ref('ko_bcr_total')}", f"={res_ref('ko_bcr_total', True)}"),
        ("Child BLL reduction during intervention period", "ug/dL", f"={res_ref('k_eff_bll_child_total')}", f"={res_ref('k_eff_bll_child_total', True)}", f"={res_ref('ko_eff_bll_child_total')}", f"={res_ref('ko_eff_bll_child_total', True)}"),
        ("Developmentally weighted child blood lead reduction used for cognition calculation", "ug/dL", f"={res_ref('k_eff_bll_child_cog_equiv')}", f"={res_ref('k_eff_bll_child_cog_equiv', True)}", f"={res_ref('ko_eff_bll_child_cog_equiv')}", f"={res_ref('ko_eff_bll_child_cog_equiv', True)}"),
        ("Child IQ gain", "IQ points", f"={res_ref('k_delta_iq')}", f"={res_ref('k_delta_iq', True)}", f"={res_ref('ko_delta_iq')}", f"={res_ref('ko_delta_iq', True)}"),
    ]
    for row in topline_rows:
        top.append(row)
    for col, width in {"A": 78, "B": 14, "C": 20, "D": 20, "E": 20, "F": 20}.items():
        top.column_dimensions[col].width = width

    for sheet in wb.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

    for sheet_name in ["Results", "Topline"]:
        sheet = wb[sheet_name]
        for row in sheet.iter_rows(min_row=3):
            for cell in row[1:3] if sheet_name == "Results" else row[2:]:
                cell.number_format = "0.00"

    wb.save(workbook_path)
    return workbook_path


def run_all(output_dir):
    output_dir.mkdir(parents=True, exist_ok=True)

    kitchen_base = simulate_kitchen(draw_params(seed=RNG_SEED))
    kohl_draws = draw_params(seed=RNG_SEED + 2)
    kohl_base = simulate_kohl(kohl_draws)
    kitchen_mode = simulate_kitchen(mode_params())
    kohl_mode = simulate_kohl(mode_params())
    kohl_girls_only, kohl_mothers_girls_trials = kohl_girls_only_robustness(kohl_draws, kohl_base)
    validate_results(kitchen_base, "kitchen")
    validate_results(kohl_base, "kohl")
    validate_results(kohl_mothers_girls_trials, "kohl_mothers_girls_only")

    export_trials(kitchen_base, output_dir, "kitchen_baseline")
    export_trials(kohl_base, output_dir, "kohl_baseline")
    export_trials(kohl_mothers_girls_trials, output_dir, "kohl_mothers_girls_only")
    kohl_girls_only.to_csv(output_dir / "kohl_mothers_girls_only_robustness.csv", index=False)

    plot_bcr_comparison(kitchen_base, kohl_base, output_dir)
    plot_single_bcr(kitchen_base, "Kitchen", output_dir)
    plot_single_bcr(kohl_base, "Kohl", output_dir)
    plot_main_pathway_figure(output_dir)

    summary_rows = []
    for package_name, df in [
        ("Kitchen Baseline", kitchen_base),
        ("Kohl Baseline", kohl_base),
    ]:
        cog = summarize_results(df, "bcr_cog")
        total = summarize_results(df, "bcr_total")
        modal_scale = modal_policy_scale_outputs(kitchen_mode if package_name == "Kitchen Baseline" else kohl_mode)
        summary_rows.append(
            {
                "package_case": package_name,
                "bcr_cog_median": cog["median"],
                "bcr_cog_p5": cog["p5"],
                "bcr_cog_p95": cog["p95"],
                "bcr_cog_fail_pct": cog["fail_pct"],
                "bcr_total_median": total["median"],
                "bcr_total_p5": total["p5"],
                "bcr_total_p95": total["p95"],
                "bcr_total_fail_pct": total["fail_pct"],
                **modal_scale,
            }
        )

    summary = pd.DataFrame(summary_rows)
    summary.to_csv(output_dir / "summary_table.csv", index=False)
    validation = pd.DataFrame(
        [
            build_validation_report("Kitchen Baseline", kitchen_base),
            build_validation_report("Kohl Baseline", kohl_base),
        ]
    )
    validation.to_csv(output_dir / "validation_checks.csv", index=False)

    kitchen_tornado = tornado_data("kitchen", metric="bcr_cog")
    kohl_tornado = tornado_data("kohl", metric="bcr_cog")
    kitchen_tornado.to_csv(output_dir / "kitchen_tornado_bcr_cog.csv", index=False)
    kohl_tornado.to_csv(output_dir / "kohl_tornado_bcr_cog.csv", index=False)
    plot_tornado(kitchen_tornado, "Kitchen", output_dir, "BCR Cog")
    plot_tornado(kohl_tornado, "Kohl", output_dir, "BCR Cog")

    nonlinear_checks = pd.DataFrame(
        [
            nonlinear_background_robustness("kitchen", other_mean=3.0, other_sd=1.5),
            nonlinear_background_robustness("kitchen", other_mean=7.0, other_sd=3.5),
            nonlinear_background_robustness("kohl", other_mean=3.0, other_sd=1.5),
            nonlinear_background_robustness("kohl", other_mean=7.0, other_sd=3.5),
        ]
    )
    nonlinear_checks.to_csv(output_dir / "nonlinear_background_robustness.csv", index=False)

    adult_spillover = pd.DataFrame(
        [
            adult_spillover_robustness("kitchen", spillover_share=0.5),
            adult_spillover_robustness("kohl", spillover_share=0.5),
        ]
    )
    adult_spillover.to_csv(output_dir / "adult_spillover_robustness.csv", index=False)

    example_nonlinear_iq_steps().to_csv(output_dir / "nonlinear_iq_examples.csv", index=False)
    export_robustness_distribution_summary(output_dir)
    export_disadvantaged_scenario(output_dir)
    export_half_effect_scenario(output_dir)
    export_modes_workbook(output_dir)
    reader_workbook = export_reader_deterministic_workbook(output_dir)
    shutil.copy2(reader_workbook, output_dir / "Simulating_sentinel_safe_products.xlsx")
    return summary


if __name__ == "__main__":
    output_dir = Path(__file__).resolve().parent / "outputs"
    summary = run_all(output_dir)
    print(summary.to_string(index=False))
